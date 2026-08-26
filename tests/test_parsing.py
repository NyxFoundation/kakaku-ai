"""スクレイパの「壊れたら黙って間違う」部分を押さえるテスト。

ネットワークは叩かない。ヒューリスティックが効いているところ
（度数分布からの中央値推定、緩い JSON、年式→世代、robots 由来の禁止パラメータ）
だけを対象にしている。
"""

from __future__ import annotations

import json

import pytest

from kakaku_ai import aggregate
from kakaku_ai.sources import carsensor, mlit, yahoo_auction
from kakaku_ai.vehicles import load_vehicles


# ------------------------------------------------------------ カーセンサー


@pytest.mark.parametrize(
    ("label", "expected"),
    [
        ("20 万円未満", (0.0, 20.0)),
        ("40 万円~", (40.0, 60.0)),
        ("400 万円~", (400.0, 420.0)),
        ("420 万円以上", (420.0, 440.0)),
    ],
)
def test_price_bin(label, expected):
    assert carsensor._parse_price_bin(label) == expected


def test_grouped_stats_interpolates_within_bin():
    # 20-40 に 1 台、40-60 に 3 台 → 中央値は 40-60 のビンの中に落ちる
    stats = carsensor._grouped_stats([((20.0, 40.0), 1), ((40.0, 60.0), 3)])
    assert stats["n"] == 4
    assert 40.0 <= stats["median"] <= 60.0
    assert stats["mean"] == pytest.approx((30 * 1 + 50 * 3) / 4)


@pytest.mark.parametrize(
    ("label", "expected"),
    [
        ("0.05 万 km未満", (0.0, 0.05)),
        ("0.5 万 | 0.7 万km", (0.5, 0.7)),
        ("10 万 | 15 万km", (10.0, 15.0)),
        ("15 万Km 以上", (15.0, 20.0)),
        ("合計", None),
    ],
)
def test_mileage_bin(label, expected):
    assert carsensor._parse_mileage_bin(label) == expected


def test_grouped_stats_empty():
    assert carsensor._grouped_stats([])["n"] == 0
    assert carsensor._grouped_stats([((0.0, 20.0), 0)])["median"] is None


def test_cross_table_column_alignment():
    """年式列と <td> が 1 対 1 で対応していること。

    ここがずれると全部の年式相場が 1 年分ずれる。実際に一度やらかした。
    """
    from bs4 import BeautifulSoup

    html = """
    <table class="defaultTable__table">
      <thead>
        <tr><th colspan="2">中古車情報の相場</th><th>2013年</th><th>2014年</th></tr>
        <tr><th colspan="2">合計 5 台</th><th>2 台</th><th>3 台</th></tr>
      </thead>
      <tbody>
        <tr><th>40 万円~</th><th>4 台</th><td>2</td><td>2</td></tr>
        <tr><th>60 万円~</th><th>1 台</th><td>&nbsp;</td><td>1</td></tr>
      </tbody>
    </table>
    """
    table = BeautifulSoup(html, "lxml").select_one("table")
    columns, rows = carsensor._parse_cross_table(table)

    assert columns == ["2013年", "2014年"]
    # 2013年 列 = 台数 2（40万円台のみ）、2014年 列 = 台数 3
    assert sum(counts[0] for _, counts in rows) == 2
    assert sum(counts[1] for _, counts in rows) == 3


# ------------------------------------------------------------------ 国交省


def test_loads_lenient_handles_trailing_commas():
    """連ラクダの CGI はテンプレート由来の末尾カンマ付き JSON を返す。"""
    raw = """
        \n\n
        { "data": [ {"a": 1}, {"b": 2}, ]
        }
        \n
    """
    assert mlit.loads_lenient(raw) == {"data": [{"a": 1}, {"b": 2}]}


def test_loads_lenient_rejects_garbage():
    with pytest.raises(ValueError):
        mlit.loads_lenient("no json here")


def test_recall_types_gathers_numbered_lists():
    row = {
        "typeList": [{"recall_type_data_car_mlit_model_name": "DBA-AGH30W"}],
        "typeList3": [{"recall_type_data_car_mlit_model_name": "DBA-AGH35W"}],
        "typeList7": [],
        "other": "無視される",
    }
    models = {t["recall_type_data_car_mlit_model_name"] for t in mlit._recall_types(row)}
    assert models == {"DBA-AGH30W", "DBA-AGH35W"}


def test_normalize_recalls_matches_by_bare_model_code():
    """排ガス記号が違っても素の型式が一致すれば拾う（3BA- と DBA- など）。"""
    vehicle = load_vehicles().by_key("alphard")
    rows = [
        {
            "recall_data_car_mlit_notification_no": "1",
            "recall_data_car_mlit_notification_date": "2025/01/22",
            "recall_data_car_mlit_defective_device": "燃料ポンプ",
            "recall_data_car_mlit_recall_car_count": "54,577",
            "typeList": [
                {
                    "recall_type_data_car_mlit_model_name": "5BA-AGH30W",
                    "recall_type_data_car_mlit_common_name": "アルファード",
                }
            ],
        },
        {
            "recall_data_car_mlit_notification_no": "2",
            "recall_data_car_mlit_defective_device": "無関係",
            "typeList": [
                {
                    "recall_type_data_car_mlit_model_name": "DBA-ZRR80W",
                    "recall_type_data_car_mlit_common_name": "ノア",
                }
            ],
        },
    ]
    out = mlit.normalize_recalls(rows, vehicle, "2026-08-22")
    assert [r["notification_no"] for r in out] == ["1"]
    assert out[0]["target_units"] == 54577


# ------------------------------------------------------------------ ヤフオク


def test_forbidden_params_are_rejected():
    """robots.txt が禁じているパラメータを付けようとしたら止まること。"""
    with pytest.raises(ValueError, match="robots.txt"):
        yahoo_auction._check_params({"auccat": 26360, "n": 100})
    yahoo_auction._check_params({"auccat": 26360, "b": 51, "p": "アルファード"})


@pytest.mark.parametrize(
    ("raw", "expected"),
    [(20240401, 202404), ("20150131", 201501), (None, None), (0, None), (12, None)],
)
def test_model_year_month(raw, expected):
    assert yahoo_auction._model_year_month(raw) == expected


def test_next_data_missing_raises():
    with pytest.raises(ValueError):
        yahoo_auction._parse_next_data("<html><body>no script</body></html>")


def test_next_data_roundtrip():
    payload = {"props": {"pageProps": {}}}
    html = f'<script id="__NEXT_DATA__" type="application/json">{json.dumps(payload)}</script>'
    assert yahoo_auction._parse_next_data(html) == payload


# ------------------------------------------------------------------ 集計


class _FakeVehicle:
    key = "test"
    name = "テスト車"

    @staticmethod
    def generation_label(_ym):
        return "1系"

    @staticmethod
    def generation_for_model_year(_year):
        return "1系"


def test_yahoo_by_year_excludes_tampered_and_repaired():
    rows = [
        {"model_year": 2020, "price": 1_000_000, "mileage_type": "REAL_MILEAGE", "repair_type": "NONE", "mileage_km": 50_000, "bid_count": 3},
        {"model_year": 2020, "price": 2_000_000, "mileage_type": "REAL_MILEAGE", "repair_type": "NONE", "mileage_km": 30_000, "bid_count": 5},
        # メーター交換 → 主系列から外れる
        {"model_year": 2020, "price": 100_000, "mileage_type": "METER_REPLACEMENT", "repair_type": "NONE", "mileage_km": 200_000, "bid_count": 1},
    ]
    out = aggregate.yahoo_by_year(rows, _FakeVehicle(), "2026-08-22")
    assert len(out) == 1
    assert out[0]["auction_n"] == 2
    assert out[0]["auction_median_manyen"] == pytest.approx(150.0)
    assert out[0]["excluded_n"] == 1
    assert out[0]["basis"] == "実走行・修復歴あり以外"


def test_yahoo_by_year_keeps_unknown_repair_history():
    """修復歴「わからない」は主系列に残し、件数だけ別に持つこと。

    個人出品では普通の申告で、落とすとサンプルが 2 割以上痩せる。
    一方で「あり」は価格が下振れするので外す。
    """
    rows = [
        {"model_year": 2018, "price": 1_000_000, "mileage_type": "REAL_MILEAGE", "repair_type": "NONE"},
        {"model_year": 2018, "price": 1_200_000, "mileage_type": "REAL_MILEAGE", "repair_type": "UNKNOWN"},
        {"model_year": 2018, "price": 300_000, "mileage_type": "REAL_MILEAGE", "repair_type": "EXISTS"},
        {"model_year": 2018, "price": 400_000, "mileage_type": "REAL_MILEAGE", "repair_type": "REPAIRED"},
    ]
    out = aggregate.yahoo_by_year(rows, _FakeVehicle(), "2026-08-22")
    assert out[0]["auction_n"] == 2          # なし + わからない
    assert out[0]["unknown_repair_n"] == 1
    assert out[0]["excluded_n"] == 2         # EXISTS / REPAIRED
    assert out[0]["auction_median_manyen"] == pytest.approx(110.0)


def test_is_usable_predicate():
    from kakaku_ai.aggregate import is_usable

    assert is_usable({"repair_type": "NONE", "mileage_type": "REAL_MILEAGE"})
    assert is_usable({"repair_type": "UNKNOWN", "mileage_type": "REAL_MILEAGE"})
    assert is_usable({})  # 何も分かっていないものは残す
    assert not is_usable({"repair_type": "EXISTS"})
    assert not is_usable({"repair_type": "REPAIRED"})
    assert not is_usable({"mileage_type": "METER_REPLACEMENT"})
    assert not is_usable({"mileage_type": "UNKNOWN_MILEAGE"})


def test_yahoo_by_year_falls_back_when_all_flagged():
    rows = [
        {"model_year": 2019, "price": 500_000, "mileage_type": "METER_REPLACEMENT", "repair_type": "NONE"},
    ]
    out = aggregate.yahoo_by_year(rows, _FakeVehicle(), "2026-08-22")
    assert out[0]["auction_n"] == 1
    assert out[0]["basis"].startswith("全件")


def test_merge_price_rows_computes_premium_and_filters_old_years():
    auction = [
        {"model_year": 2012, "auction_median_manyen": 50.0, "auction_n": 3, "generation": "1系"},
        {"model_year": 2020, "auction_median_manyen": 100.0, "auction_n": 4, "generation": "1系"},
    ]
    retail = [
        {"model_year": 2020, "retail_median_manyen": 150.0, "listing_count": 20, "generation": "1系"},
        {"model_year": 2021, "retail_median_manyen": 180.0, "listing_count": 10, "generation": "1系"},
    ]
    out = aggregate.merge_price_rows(
        auction, retail, _FakeVehicle(), "2026-08-22", model_year_from=2013
    )
    years = [r["model_year"] for r in out]
    assert years == [2020, 2021]  # 2012 は対象外

    row2020 = out[0]
    assert row2020["retail_minus_auction_manyen"] == pytest.approx(50.0)
    assert row2020["retail_premium_pct"] == pytest.approx(50.0)

    # 片側しかない年は乖離を出さない
    assert out[1]["retail_premium_pct"] is None


# ------------------------------------------------------------ 車種マスタ


def test_generation_resolution_at_boundaries():
    alphard = load_vehicles().by_key("alphard")
    assert alphard.generation_label(201412) == "20系"
    assert alphard.generation_label(201501) == "30系前期"  # 境界は from 側に含む
    assert alphard.generation_label(201711) == "30系前期"
    assert alphard.generation_label(201712) == "30系後期"
    assert alphard.generation_label(202306) == "40系"
    assert alphard.generation_label(None) == ""


def test_every_vehicle_has_ids():
    for vehicle in load_vehicles():
        assert vehicle.generations, f"{vehicle.name}: 世代が未設定"
        assert vehicle.all_models, f"{vehicle.name}: 型式が未設定"
        assert vehicle.carsensor_codes, f"{vehicle.name}: カーセンサーのコードが引けていない"
        assert vehicle.minkara_slug, f"{vehicle.name}: みんカラの slug が未設定"


# ------------------------------------------------------------------ みんカラ


def test_review_summary_dedupes_identical_text():
    """別レビューでも本文が丸かぶりのことがある（みんカラに実在）。

    URL が違うので個票としては別物だが、代表コメントに同じ文を並べても意味がない。
    """
    from kakaku_ai.sources import minkara

    rows = [
        {"model_year": 2008, "score_overall": 4, "generation": "50系",
         "good_points": "車内広く使える", "bad_points": ""},
        {"model_year": 2008, "score_overall": 5, "generation": "50系",
         "good_points": "車内広く使える", "bad_points": ""},
        {"model_year": 2008, "score_overall": 3, "generation": "50系",
         "good_points": "3列目シートが床下収納になる", "bad_points": "燃費が悪い"},
    ]
    out = minkara.summarize(rows, _FakeVehicle(), "2026-08-22")
    assert len(out) == 1
    assert out[0]["review_count"] == 3
    assert out[0]["good_points"] == "車内広く使える / 3列目シートが床下収納になる"
    assert out[0]["bad_points"] == "燃費が悪い"
    assert out[0]["score_overall"] == pytest.approx(4.0)


# -------------------------------------------------------------------- グラフ


def test_year_chart_spans_missing_years():
    """落札のない年式でも折れ線が途切れないこと。

    openpyxl の既定は display_blanks='gap' で、空セルのところで線が切れる。
    年式別の落札は歯抜けになりがちなので 'span' でまたぐ。
    """
    from openpyxl import Workbook

    from kakaku_ai import charts

    listings = [
        {"vehicle_name": "テスト車", "model_year": y, "price": p,
         "mileage_type": "REAL_MILEAGE", "repair_type": "NONE", "mileage_km": 50_000}
        # 2015 と 2017 は落札なし = 歯抜け
        for y, p in [(2013, 500_000), (2014, 700_000), (2016, 1_200_000), (2018, 2_000_000)]
    ]
    wb = Workbook()
    ws = wb.create_sheet("グラフ_年式別")
    charts.year_sheet(ws, listings, 2013)

    assert len(ws._charts) == 1
    assert ws._charts[0].display_blanks == "span"

    # 歯抜けの年式は「0」ではなく空セルのまま（0 だと谷があるように見えてしまう）
    header = [c.value for c in ws[5]]
    assert header[1:] == ["2013年", "2014年", "2016年", "2018年"]


def test_pivot_chart_spans_missing_snapshots(tmp_path, monkeypatch):
    """取得できなかった週があっても推移の折れ線が途切れないこと。"""
    from openpyxl import Workbook

    from kakaku_ai import excel

    wb = Workbook()
    ws = wb.active
    rows = [
        {"snapshot_date": "2026-08-01", "vehicle_name": "テスト車", "model_year": 2020,
         "generation": "1系", "auction_median_manyen": 100.0},
        # 2026-08-08 は欠測
        {"snapshot_date": "2026-08-15", "vehicle_name": "テスト車", "model_year": 2020,
         "generation": "1系", "auction_median_manyen": 90.0},
    ]
    excel._pivot(ws, rows, ["2026-08-01", "2026-08-08", "2026-08-15"],
                 "auction_median_manyen", "落札中央値")

    assert len(ws._charts) == 1
    assert ws._charts[0].display_blanks == "span"


# ------------------------------------------------------------------ 累計プール


def test_pooled_listings_dedupe_and_prefer_richer(tmp_path, monkeypatch):
    """週をまたいで同じ落札が出てきても 1 件に名寄せし、情報量の多い側を残すこと。

    ヤフオクの落札検索は終了180日ぶんしか返さないので、週次の窓は重なる。
    ここで重複排除できないと同じ車を何度も数えてしまう。
    """
    import json

    from kakaku_ai import store

    monkeypatch.setattr(store, "SNAPSHOT_DIR", tmp_path)

    def put(snap, rows):
        d = tmp_path / snap
        d.mkdir(parents=True, exist_ok=True)
        (d / "auction_listings.jsonl").write_text(
            "\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n",
            encoding="utf-8",
        )

    put("2026-08-01", [
        {"auction_id": "a1", "snapshot_date": "2026-08-01", "price": 100, "end_time": "2026-07-30T00:00:00+09:00"},
        {"auction_id": "a2", "snapshot_date": "2026-08-01", "price": 200, "end_time": "2026-07-31T00:00:00+09:00"},
    ])
    # a1 は再登場（今度は年式つき）、a3 は新規
    put("2026-08-08", [
        {"auction_id": "a1", "snapshot_date": "2026-08-08", "price": 100, "model_year": 2018,
         "mileage_km": 50_000, "end_time": "2026-07-30T00:00:00+09:00"},
        {"auction_id": "a3", "snapshot_date": "2026-08-08", "price": 300, "end_time": "2026-08-05T00:00:00+09:00"},
    ])

    pool = store.pooled_auction_listings()
    assert len(pool) == 3
    by_id = {r["auction_id"]: r for r in pool}
    # 情報量の多い 2 週目の版が残る
    assert by_id["a1"]["model_year"] == 2018
    # 初出は 1 週目のまま
    assert by_id["a1"]["first_seen_snapshot"] == "2026-08-01"
    assert by_id["a3"]["first_seen_snapshot"] == "2026-08-08"
    # 終了日時で並ぶ
    assert [r["auction_id"] for r in pool] == ["a1", "a2", "a3"]


def test_detail_apply_fills_only_missing_fields():
    """検索結果で取れている値は上書きせず、欠けているところだけ詳細で埋めること。"""
    from kakaku_ai.sources import yahoo_detail

    listings = [
        {"auction_id": "x1", "model_year": 2020, "mileage_km": 30_000,
         "mileage_type": "REAL_MILEAGE", "repair_type": "NONE"},
        {"auction_id": "x2"},
    ]
    store_data = {
        "x1": {"auction_id": "x1", "first_reg_year": 1999, "mileage_km": 999_999,
               "grade": "Z", "fetched_from": "detail_page"},
        "x2": {"auction_id": "x2", "first_reg_year": 2015, "first_reg_month": 3,
               "mileage_km": 80_000, "mileage_status": "メーター交換歴あり",
               "repair_history": "あり", "grade": "G", "fetched_from": "detail_page"},
    }
    yahoo_detail.apply_to(listings, store_data)

    # 既にある値は詳細で上書きされない
    assert listings[0]["model_year"] == 2020
    assert listings[0]["mileage_km"] == 30_000
    # 詳細にしかない項目は足される
    assert listings[0]["grade"] == "Z"

    # 欠けていたほうは埋まる
    assert listings[1]["model_year"] == 2015
    assert listings[1]["model_year_month"] == 201503
    assert listings[1]["year_source"] == "detail_page"
    assert listings[1]["mileage_type"] == "METER_REPLACEMENT"
    assert listings[1]["repair_type"] == "REPAIRED"


# ---------------------------------------------------------------- ジモティー


def test_jmty_dedupes_repeat_postings():
    """同じ車の多重投稿を 1 件に落とすこと。

    ジモティーには 1 台を文言だけ変えて何度も出す業者がいる
    （実データでウィッシュの 1 台が 8 投稿）。
    """
    from kakaku_ai.sources import jmty

    rows = [
        {"asking_price": 1_192_000, "model_year": 2013, "mileage_km": 90_000, "title": "A"},
        {"asking_price": 1_192_000, "model_year": 2013, "mileage_km": 90_000, "title": "B"},
        {"asking_price": 1_192_000, "model_year": 2013, "mileage_km": 90_000, "title": "C"},
        {"asking_price": 770_000, "model_year": 2014, "mileage_km": 66_000, "title": "D"},
        # 年式・距離が未記入のものは同一判定できないので残す
        {"asking_price": 500_000, "model_year": None, "mileage_km": None, "title": "E"},
        {"asking_price": 500_000, "model_year": None, "mileage_km": None, "title": "F"},
    ]
    out = jmty._dedupe(rows)
    assert [r["title"] for r in out] == ["A", "D", "E", "F"]


def test_jmty_keyword_search_checks_title():
    """キーワード検索は本文にも当たるので、タイトルで車種を確かめること。

    「プリウスα」で検索してヴォクシーが混ざった実例がある。
    """
    from kakaku_ai.sources import jmty
    from kakaku_ai.vehicles import load_vehicles

    vehicle = load_vehicles().by_key("prius_alpha")
    ok = {"title": "H25 プリウスα ５人乗り(ＨＶバッテリー交換済み)"}
    ng = {"title": "【総額65万円】車検長期ヴォクシーHV/低燃費ハイブリッドミニバン"}
    plain_prius = {"title": "30プリウス 後期 フルエアロ 90000キロ【美車】"}

    assert jmty._matches_vehicle(ok, vehicle, has_category=False)
    assert not jmty._matches_vehicle(ng, vehicle, has_category=False)
    assert not jmty._matches_vehicle(plain_prius, vehicle, has_category=False)
    # タイトルパターンが指定されていればカテゴリ指定でも照合する。
    # jmty のカテゴリは車種より粗いことがあり、三菱の「デリカ」には
    # D:5 / D:2 / ミニ が同居していて、カテゴリを信じると別車種が混ざる。
    assert not jmty._matches_vehicle(ng, vehicle, has_category=True)


def test_jmty_category_without_pattern_is_trusted():
    """パターン未指定ならカテゴリを信じる（アルファード等）。"""
    from kakaku_ai.sources import jmty
    from kakaku_ai.vehicles import load_vehicles

    vehicle = load_vehicles().by_key("alphard")
    assert vehicle.jmty_title_pattern is None
    assert jmty._matches_vehicle({"title": "何でもよい"}, vehicle, has_category=True)


def test_jmty_delica_pattern_separates_d5_from_d2():
    """デリカは D:5 だけを拾い、D:2 / ミニ を弾くこと。"""
    import re

    from kakaku_ai.vehicles import load_vehicles

    pat = load_vehicles().by_key("delica_d5").jmty_title_pattern
    assert all(re.search(pat, t) for t in
               ["即乗りＯＫ デリカD5 Dパワーパッケージ", "三菱 デリカD:5 2.2", "デリカＤ５ ジャスパー"])
    assert not any(re.search(pat, t) for t in
                   ["三菱 デリカＤ：２ Ｘ", "三菱 デリカミニ", "デリカD2 1.2"])


def test_jmty_by_year_counts_direct_posts_separately():
    """業者・個人は分けずに集計し、直接投稿の件数だけ内訳として持つこと。"""
    from kakaku_ai.sources import jmty

    rows = [
        {"model_year": 2018, "asking_price": 2_000_000, "is_alliance": True},
        {"model_year": 2018, "asking_price": 2_400_000, "is_alliance": True},
        {"model_year": 2018, "asking_price": 2_800_000, "is_alliance": False},
        {"model_year": None, "asking_price": 999_000, "is_alliance": False},
    ]
    out = jmty.by_year(rows, _FakeVehicle(), "2026-08-23")
    assert len(out) == 1
    assert out[0]["jmty_n"] == 3
    assert out[0]["jmty_direct_n"] == 1
    assert out[0]["jmty_median_manyen"] == pytest.approx(240.0)


def test_generation_for_model_year_handles_mid_year_changes():
    """年の途中でモデルチェンジした年式を取りこぼさないこと。

    月が分からない年式を「6月とみなす」だけだと、7月開始の世代（シエンタ170系）や
    10月開始の世代（エスクァイア80系）が世代なしになる。実データでシエンタ2015が
    落札12件・掲載378件ありながら世代不明で出ていた。
    """
    vs = load_vehicles()

    sienta = vs.by_key("sienta")
    assert sienta.generation_for_model_year(2014) == "80系"
    assert sienta.generation_for_model_year(2015) == "80系/170系"  # 7月に交代
    assert sienta.generation_for_model_year(2016) == "170系"

    esquire = vs.by_key("esquire")
    assert esquire.generation_for_model_year(2014) == "80系"  # 10月発売でも拾う

    alphard = vs.by_key("alphard")
    assert alphard.generation_for_model_year(2023) == "30系後期/40系"  # 6月に交代
    assert alphard.generation_for_model_year(2024) == "40系"

    isis = vs.by_key("isis")
    assert isis.generation_for_model_year(2017) == "10系"  # 12月終了

    # 生産終了後の年式は該当なし（データの揺れ）
    assert vs.by_key("prius_alpha").generation_for_model_year(2022) == ""
    assert alphard.generation_for_model_year(None) == ""


# ------------------------------------------------------- 店頭在庫の個体追跡


def test_stock_tracking_marks_disappeared_listings(tmp_path, monkeypatch):
    """先週まであった掲載が今週消えたら「売れた」印をつけること。

    店頭の成約価格はどこも公開していないので、これが唯一の手がかりになる。
    """
    from kakaku_ai.sources import carsensor_listings as cl

    monkeypatch.setattr(cl, "STORE_PATH", tmp_path / "stock.jsonl")

    class _V:
        key = "test"
        name = "テスト車"
        carsensor_codes = ("TO_S001",)

        @staticmethod
        def generation_for_model_year(_y):
            return "1系"

    week1 = [
        {"listing_id": "AU1", "vehicle_key": "test", "vehicle_name": "テスト車",
         "carsensor_code": "TO_S001", "model_year": 2018, "generation": "1系",
         "total_price_manyen": 300.0, "base_price_manyen": 290.0, "mileage_km": 50000,
         "inspection": "", "repair_history": "なし", "warranty": "", "url": "u1"},
        {"listing_id": "AU2", "vehicle_key": "test", "vehicle_name": "テスト車",
         "carsensor_code": "TO_S001", "model_year": 2018, "generation": "1系",
         "total_price_manyen": 320.0, "base_price_manyen": 310.0, "mileage_km": 40000,
         "inspection": "", "repair_history": "なし", "warranty": "", "url": "u2"},
    ]
    # 2週目: AU1 は値下げして残り、AU2 は消える
    week2 = [dict(week1[0], total_price_manyen=285.0)]

    def fake_fetch(rows):
        return lambda fetcher, vehicle, code, year: rows if year == 2018 else []

    monkeypatch.setattr(cl, "_fetch_year", fake_fetch(week1))
    c1 = cl.track(None, [_V()], "2026-08-23", model_year_from=2018, model_year_to=2018)
    assert c1 == {"tracked": 2, "seen_this_week": 2, "newly_delisted": 0}
    assert cl.delisted_rows("2026-08-23") == []

    monkeypatch.setattr(cl, "_fetch_year", fake_fetch(week2))
    c2 = cl.track(None, [_V()], "2026-08-30", model_year_from=2018, model_year_to=2018)
    assert c2["newly_delisted"] == 1

    gone = cl.delisted_rows("2026-08-30")
    assert len(gone) == 1
    assert gone[0]["listing_id"] == "AU2"
    assert gone[0]["last_price_manyen"] == pytest.approx(320.0)
    assert gone[0]["delisted_on"] == "2026-08-30"

    # 残ったほうは値下げが履歴に残る
    store = cl.load_store()
    assert store["AU1"]["price_history"] == [["2026-08-23", 300.0], ["2026-08-30", 285.0]]
    assert store["AU1"]["delisted_on"] is None


def test_stock_tracking_undoes_delisting_on_relist(tmp_path, monkeypatch):
    """一度消えた掲載が戻ってきたら「売れた」を取り消すこと。"""
    from kakaku_ai.sources import carsensor_listings as cl

    monkeypatch.setattr(cl, "STORE_PATH", tmp_path / "stock.jsonl")

    class _V:
        key = "test"
        name = "テスト車"
        carsensor_codes = ("TO_S001",)

        @staticmethod
        def generation_for_model_year(_y):
            return "1系"

    row = {"listing_id": "AU9", "vehicle_key": "test", "vehicle_name": "テスト車",
           "carsensor_code": "TO_S001", "model_year": 2019, "generation": "1系",
           "total_price_manyen": 200.0, "base_price_manyen": 195.0, "mileage_km": 60000,
           "inspection": "", "repair_history": "なし", "warranty": "", "url": "u9"}

    monkeypatch.setattr(cl, "_fetch_year", lambda f, v, c, y: [row] if y == 2019 else [])
    cl.track(None, [_V()], "2026-08-23", model_year_from=2019, model_year_to=2019)

    monkeypatch.setattr(cl, "_fetch_year", lambda f, v, c, y: [])
    cl.track(None, [_V()], "2026-08-30", model_year_from=2019, model_year_to=2019)
    assert len(cl.delisted_rows("2026-08-30")) == 1

    monkeypatch.setattr(cl, "_fetch_year", lambda f, v, c, y: [row] if y == 2019 else [])
    cl.track(None, [_V()], "2026-09-06", model_year_from=2019, model_year_to=2019)
    assert cl.delisted_rows("2026-09-06") == []


def test_partial_run_keeps_untouched_datasets(tmp_path, monkeypatch):
    """--sources で絞って回しても、回さなかったデータセットを消さないこと。

    実際に `--sources stock` だけ回してその日のスナップショットを
    全部 0 にしてしまったことがある。
    """
    import json

    from kakaku_ai import pipeline, store

    monkeypatch.setattr(store, "SNAPSHOT_DIR", tmp_path)
    snap = "2026-08-23"
    d = tmp_path / snap
    d.mkdir(parents=True)
    (d / "auction_listings.jsonl").write_text(
        json.dumps({"auction_id": "a1", "vehicle_key": "alphard"}) + "\n", encoding="utf-8"
    )

    # 何も作らなかったデータセットは既存が残る
    kept = pipeline._merge_with_existing(snap, "auction_listings", [], None)
    assert [r["auction_id"] for r in kept] == ["a1"]

    # 作ったなら差し替わる
    fresh = pipeline._merge_with_existing(
        snap, "auction_listings", [{"auction_id": "a2", "vehicle_key": "alphard"}], None
    )
    assert [r["auction_id"] for r in fresh] == ["a2"]


def test_partial_run_with_only_keeps_other_vehicles(tmp_path, monkeypatch):
    """--only で車種を絞ったとき、対象外の車種の行を消さないこと。"""
    import json

    from kakaku_ai import pipeline, store

    monkeypatch.setattr(store, "SNAPSHOT_DIR", tmp_path)
    snap = "2026-08-23"
    d = tmp_path / snap
    d.mkdir(parents=True)
    (d / "price_by_year.jsonl").write_text(
        "\n".join(
            json.dumps(r)
            for r in [
                {"vehicle_key": "alphard", "model_year": 2018},
                {"vehicle_key": "noah", "model_year": 2018},
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    merged = pipeline._merge_with_existing(
        snap, "price_by_year", [{"vehicle_key": "alphard", "model_year": 2019}], ["alphard"]
    )
    keys = sorted((r["vehicle_key"], r["model_year"]) for r in merged)
    # ノアは残り、アルファードだけ差し替わる
    assert keys == [("alphard", 2019), ("noah", 2018)]


def test_yahoo_search_warns_when_truncated(caplog, monkeypatch):
    """ページ上限で打ち切ったら黙らずに警告すること。

    黙って切ると「全部取れている」と思ったまま相場を出してしまう。
    """
    import json
    import logging

    from kakaku_ai.sources import yahoo_auction as ya

    total = ya.MAX_PAGES * ya.PAGE_SIZE + 500

    class _F:
        def get_text(self, url, params=None):
            return (
                '<script id="__NEXT_DATA__">'
                + json.dumps({"props": {"pageProps": {"initialState": {"search": {"items": {
                    "listing": {
                        "totalResultsAvailable": total,
                        "items": [{"auctionId": f"x{i}"} for i in range(ya.PAGE_SIZE)],
                    }}}}}}})
                + "</script>"
            )

    with caplog.at_level(logging.WARNING):
        got = list(ya._search(_F(), {"auccat": 26360, "p": "テスト"}))

    assert len(got) == ya.MAX_PAGES * ya.PAGE_SIZE
    assert any("打ち切った" in r.getMessage() for r in caplog.records)


# ------------------------------------------------------- 出品中オークション監視


def test_open_search_uses_brand_id_not_auccat():
    """出品中の検索は brand_id を使うこと。

    /carsearch?auccat=... は auccat を黙って無視して全車種59,000件を返す。
    これに気づかず組んだせいで、セレナのつもりで全車種の先頭6ページを拾い、
    同じ出品が複数車種に重複して出ていた。
    """
    import inspect

    from kakaku_ai.sources import yahoo_open

    src = inspect.getsource(yahoo_open._search)
    assert '"brand_id"' in src
    assert '"auccat"' not in src


def test_price_judgement_ignores_pre_bidding_price():
    """競り上がる前の現在価格では相場判定しないこと。

    ¥1スタートの新規出品が「相場より99%安い」として 300件中220件ヒットし、
    通知が使い物にならなかった。判定してよいのは即決価格か終了間際だけ。
    """
    from datetime import datetime, timedelta, timezone

    from kakaku_ai.watch import _judgeable_price

    now = datetime(2026, 8, 24, 12, 0, tzinfo=timezone.utc)
    far = (now + timedelta(days=5)).isoformat()
    soon = (now + timedelta(hours=3)).isoformat()

    # 出品直後で入札前 → 判定しない
    price, kind, _ = _judgeable_price(
        {"current_price": 1000, "end_time": far, "buy_now_price": None}, now
    )
    assert price is None and "保留" in kind

    # 即決があるなら、いつでも判定してよい
    price, kind, _ = _judgeable_price(
        {"current_price": 1000, "end_time": far, "buy_now_price": 1_800_000}, now
    )
    assert price == 180.0 and kind == "即決"

    # 終了間際なら現在価格で判定してよい
    price, kind, _ = _judgeable_price(
        {"current_price": 1_500_000, "end_time": soon, "buy_now_price": None}, now
    )
    assert price == 150.0 and kind.startswith("終了")


def test_threshold_widens_when_year_has_few_samples():
    """実績の薄い年式では「相場より安い」と言い切らないこと。

    年式ごとの落札件数はまちまちで、n=3 の年式と n=83 の年式を同じしきい値で
    扱うとノイズを掘り出し物として流してしまう。
    """
    from kakaku_ai.watch import CHEAP_PCT, PriceModel

    m = PriceModel.__new__(PriceModel)
    m._count = {2017: 83, 2019: 9, 2023: 3}

    assert m.slack(2017) == 1.0
    assert m.slack(2019) > m.slack(2017)
    assert m.slack(2023) > m.slack(2019)
    assert CHEAP_PCT * m.slack(2023) < CHEAP_PCT  # より大きな乖離を要求する


def test_price_model_uses_per_year_median_not_a_line():
    """年式は実績の中央値をそのまま使い、直線で外挿しないこと。

    log(価格)~年式 の直線を張ったら、学習データ自身での残差が
    2013年 -21.6% 〜 2018年 +51.3% と大きく歪み、出品のほぼ全部が
    「相場より高い」と判定された。
    """
    from kakaku_ai.watch import PriceModel

    rows = []
    for year, price in ((2015, 60.0), (2016, 80.0), (2017, 200.0)):  # 直線では表せない上がり方
        for i in range(4):
            rows.append({"model_year": year, "mileage_km": 100_000, "price": price * 10_000,
                         "mileage_type": "REAL_MILEAGE", "repair_type": "NONE"})
    m = PriceModel("t", rows)

    for year, price in ((2015, 60.0), (2016, 80.0), (2017, 200.0)):
        assert m.expected_manyen(year, 100_000) == pytest.approx(price, rel=0.01)
    # 実績の無い年式は判定しない（外挿しない）
    assert m.expected_manyen(2020, 100_000) is None


def test_buy_now_is_compared_against_buy_now():
    """即決は落札想定ではなく即決どうしで比べること。

    即決は実測で落札想定より一律 +33% 高い（今すぐ買える確実性への上乗せ）。
    そのまま比べると即決が全部「相場より高い」になり、
    実際に通知が「相場より XX% 高い」ばかりになった。
    """
    from kakaku_ai.watch import _rebase_buy_now

    rows = [{"judge_kind": "即決", "vehicle_key": "v", "deviation_pct": d} for d in
            (10.0, 30.0, 33.0, 36.0, 60.0)]
    rows.append({"judge_kind": "終了3h前", "vehicle_key": "v", "deviation_pct": 5.0})
    _rebase_buy_now(rows)

    # 即決の中央値（33%）が基準になり、そこからの差に置き換わる
    assert rows[0]["deviation_pct"] == pytest.approx(-23.0)
    assert rows[2]["deviation_pct"] == pytest.approx(0.0)
    assert rows[0]["benchmark"] == "即決どうし"
    assert rows[0]["deviation_vs_hammer_pct"] == 10.0
    # 終了間際は落札想定のまま
    assert rows[-1]["deviation_pct"] == 5.0
    assert rows[-1]["benchmark"] == "落札想定"


def test_pick_requires_a_judgeable_price():
    """価格を判定できない出品は通知しないこと。"""
    from kakaku_ai.watch import pick

    rows = [
        {"auction_id": "a", "deviation_pct": None, "risk_strong": ["修復歴あり"],
         "cheap_threshold_pct": -25.0, "pricey_threshold_pct": 30.0, "seller_is_store": False},
        {"auction_id": "b", "deviation_pct": -40.0, "risk_strong": [],
         "cheap_threshold_pct": -25.0, "pricey_threshold_pct": 30.0, "seller_is_store": False},
    ]
    assert [r["auction_id"] for r in pick(rows)] == ["b"]


def test_repair_filter_excludes_unknown_as_well_as_exists():
    """修復歴「なし」で絞るとき、「わからない」は通さないこと。

    「なし」と「わからない」は買い手にとって別物。落札実績では 2,375件中
    416件（17%）が「わからない」で、これを「なし」に含めると意味が変わる。
    """
    from kakaku_ai.watch import repair_ok

    assert repair_ok({"repair_type": "NONE"}, "none")
    assert not repair_ok({"repair_type": "EXISTS"}, "none")
    assert not repair_ok({"repair_type": "REPAIRED"}, "none")
    assert not repair_ok({"repair_type": "UNKNOWN"}, "none")

    # 一覧に記載が無い段階では保留して残し、商品ページで確認したあとに落とす
    assert repair_ok({}, "none", resolved=False)
    assert not repair_ok({}, "none", resolved=True)

    # any なら何も落とさない
    assert all(repair_ok({"repair_type": v}, "any") for v in ("NONE", "EXISTS", "UNKNOWN", None))


def test_pick_applies_repair_filter():
    from kakaku_ai.watch import pick

    base = {"deviation_pct": -40.0, "cheap_threshold_pct": -25.0,
            "pricey_threshold_pct": 30.0, "seller_is_store": False, "risk_strong": []}
    rows = [
        {**base, "auction_id": "ok", "repair_type": "NONE"},
        {**base, "auction_id": "ng", "repair_type": "EXISTS"},
        {**base, "auction_id": "unk", "repair_type": "UNKNOWN"},
    ]
    assert [r["auction_id"] for r in pick(rows, repair="none")] == ["ok"]
    assert len(pick(rows, repair="any")) == 3


def test_rebase_buy_now_is_not_applied_twice():
    """即決の基準を二度掛けしないこと。

    詳細取得のあとに一部だけ評価し直したら、その3件から基準を計算し直して
    通知の表示と抽出条件が食い違った（「相場より11%高い」が安い側として出た）。
    """
    from kakaku_ai.watch import _rebase_buy_now

    rows = [{"judge_kind": "即決", "vehicle_key": "v", "deviation_pct": d} for d in (10.0, 30.0, 50.0)]
    _rebase_buy_now(rows)
    once = [r["deviation_pct"] for r in rows]
    _rebase_buy_now(rows)  # 二度目は何もしない
    assert [r["deviation_pct"] for r in rows] == once


# ---------------------------------------------------------------- 既読管理


def test_seen_list_prunes_only_finished_auctions(tmp_path, monkeypatch):
    """終了して十分に経った出品だけ既読から落とすこと。

    まだ終わっていないものを落とすと再通知してしまう。
    """
    from datetime import datetime, timedelta, timezone

    from kakaku_ai import notify

    monkeypatch.setattr(notify, "SEEN_PATH", tmp_path / "seen.json")
    now = datetime.now(timezone.utc)
    seen = {
        "live": (now + timedelta(days=3)).isoformat(),      # まだ出品中
        "recent": (now - timedelta(days=2)).isoformat(),    # 終了直後
        "old": (now - timedelta(days=60)).isoformat(),      # とっくに終了
        "unknown": "",                                       # 終了日時が分からない
    }
    notify.save_seen(seen)
    kept = notify.load_seen()
    assert set(kept) == {"live", "recent", "unknown"}


def test_seen_list_reads_old_list_format(tmp_path, monkeypatch):
    """旧形式（ただのIDリスト）も読めること。"""
    import json

    from kakaku_ai import notify

    path = tmp_path / "seen.json"
    path.write_text(json.dumps({"auction_ids": ["a", "b"]}), encoding="utf-8")
    monkeypatch.setattr(notify, "SEEN_PATH", path)
    assert set(notify.load_seen()) == {"a", "b"}


def test_channel_history_extracts_auction_ids(monkeypatch):
    """Slack の履歴から auction_id を拾えること。

    ローカルの既読ファイルを消してしまい、投稿済み64件のうち28件が
    既読から抜けたことがある。履歴が拾えれば、それでも再通知しない。
    """
    from kakaku_ai import notify

    class _Resp:
        @staticmethod
        def json():
            return {
                "ok": True,
                "messages": [
                    {"attachments": [{"blocks": [{"text": {
                        "text": "<https://page.auctions.yahoo.co.jp/jp/auction/k1241772531|車>"}}]}]},
                    {"text": "https://page.auctions.yahoo.co.jp/jp/auction/c1240157168"},
                ],
                "response_metadata": {},
            }

    monkeypatch.setattr(notify, "_token", lambda: "x")
    monkeypatch.setattr(notify.requests, "get", lambda *a, **k: _Resp())
    assert notify.posted_in_channel("C") == {"k1241772531", "c1240157168"}


def test_channel_history_failure_does_not_block(monkeypatch):
    """履歴が読めなくても通知そのものは止めないこと。"""
    from kakaku_ai import notify

    monkeypatch.setattr(notify, "_token", lambda: "x")

    def boom(*a, **k):
        raise RuntimeError("network down")

    monkeypatch.setattr(notify.requests, "get", boom)
    assert notify.posted_in_channel("C") == set()


# ---------------------------------------------------------------------- 車検


def test_inspection_date_normalises_both_formats():
    """車検の表記がソースで違うので、西暦の年月に寄せること。

    ヤフオク商品ページは "R10/04"（令和）、カーセンサー在庫は "2028(R10)年12月"。
    「車検整備付」のように日付が無い表記もそのまま来る。
    """
    from kakaku_ai.aggregate import inspection_year_month

    assert inspection_year_month("R10/04") == 202804   # 令和10年 = 2028年
    assert inspection_year_month("R8/11") == 202611
    assert inspection_year_month("2028(R10)年12月") == 202812
    assert inspection_year_month("2027(R09)年01月") == 202701
    # 日付が無いものは None（表示用の文字列は別に残す）
    for v in ("車検整備付", "車検整備無", "", None):
        assert inspection_year_month(v) is None


def test_months_until_handles_expired():
    from datetime import date

    from kakaku_ai.aggregate import months_until

    today = date(2026, 8, 26)
    assert months_until(202804, today) == 20
    assert months_until(202611, today) == 3
    assert months_until(202606, today) == -2  # 切れている
    assert months_until(None, today) is None
