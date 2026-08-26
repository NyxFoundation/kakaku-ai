"""旧車ブックのグラフシート。

表だけ渡されても 1 万台は読めないので、「どのメーカーに何が残っているか」
「年式ごとに何台あって、いくらか」「店頭とヤフオクでどれだけ違うか」を
先に絵で出す。表は絵で当たりを付けたあとに絞り込む用。

4 枚:

* `グラフ_メーカー別` … メーカーごとの在庫台数と価格中央値
* `グラフ_年式別`   … 年式ごとの台数と価格中央値（1988〜2001）
* `グラフ_車種別`   … 在庫の多い車種の価格帯（最安・中央値・最高）
* `グラフ_落札比較` … 同じ車種の 店頭価格 と ヤフオク落札価格 の差
"""

from __future__ import annotations

import logging
import statistics as st
from collections import defaultdict
from typing import Any

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.worksheet import Worksheet

from .charts import BORDER, INT_FMT, MANYEN_FMT, _header, _title

log = logging.getLogger(__name__)

MANYEN = 10_000
TOP_MAKERS = 18
TOP_MODELS = 20
MIN_AUCTION_SAMPLES = 3  # これ未満の車種は落札比較に載せない


def _price(row: dict[str, Any]) -> float | None:
    return row.get("total_price_manyen") or row.get("base_price_manyen")


def _rows(ws: Worksheet, start: int, records: list[tuple], formats: dict[int, str]) -> int:
    for offset, record in enumerate(records):
        for col_index, value in enumerate(record, start=1):
            cell = ws.cell(row=start + 1 + offset, column=col_index, value=value)
            cell.border = BORDER
            if col_index in formats:
                cell.number_format = formats[col_index]
    return start + len(records)


# ----------------------------------------------------------- 1. メーカー別


def maker_sheet(ws: Worksheet, listings: list[dict[str, Any]]) -> None:
    _title(
        ws,
        "メーカー別 在庫台数と価格",
        "1988〜2001年式でカーセンサーに出ている台数。価格は支払総額の中央値。"
        "台数が少ないメーカーは、そもそも玉が無いということ。",
    )

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in listings:
        groups[row.get("maker") or "不明"].append(row)

    records = []
    for maker, group in groups.items():
        prices = [p for p in (_price(r) for r in group) if p]
        records.append((maker, len(group), round(st.median(prices), 1) if prices else None))
    records.sort(key=lambda r: -r[1])
    records = records[:TOP_MAKERS]

    start = 4
    _header(ws, start, ["メーカー", "在庫台数", "価格中央値(万円)"])
    last = _rows(ws, start, records, {2: INT_FMT, 3: MANYEN_FMT})

    chart = BarChart()
    chart.type = "bar"
    chart.title = f"メーカー別 在庫台数（上位{len(records)}）"
    chart.x_axis.title = "台数"
    chart.height, chart.width = 14, 20
    chart.add_data(Reference(ws, min_col=2, max_col=2, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "F4")

    price = BarChart()
    price.type = "bar"
    price.title = "メーカー別 価格中央値（万円）"
    price.x_axis.title = "万円"
    price.height, price.width = 14, 20
    price.add_data(Reference(ws, min_col=3, max_col=3, min_row=start, max_row=last),
                   titles_from_data=True)
    price.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(price, "R4")

    ws.column_dimensions["A"].width = 20
    for col in "BC":
        ws.column_dimensions[col].width = 16


# ------------------------------------------------------------- 2. 年式別


def year_sheet(ws: Worksheet, listings: list[dict[str, Any]],
               year_from: int, year_to: int) -> None:
    _title(
        ws,
        "年式別 在庫台数と価格",
        "古いほど台数が減り、価格は上がる（残っている個体が選別されるため）。"
        "台数の谷は「そもそも探しても出てこない年式」を意味する。",
    )

    by_year: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in listings:
        if row.get("model_year"):
            by_year[row["model_year"]].append(row)

    records = []
    for year in range(year_from, year_to + 1):
        group = by_year.get(year, [])
        prices = [p for p in (_price(r) for r in group) if p]
        mileages = [r["mileage_km"] for r in group if r.get("mileage_km")]
        records.append((
            year,
            len(group),
            round(st.median(prices), 1) if prices else None,
            round(st.median(mileages) / MANYEN, 1) if mileages else None,
        ))

    start = 4
    _header(ws, start, ["年式", "在庫台数", "価格中央値(万円)", "走行中央値(万km)"])
    last = _rows(ws, start, records, {2: INT_FMT, 3: MANYEN_FMT, 4: "0.0"})

    chart = BarChart()
    chart.type = "col"
    chart.title = f"年式別 在庫台数（{year_from}〜{year_to}）"
    chart.y_axis.title = "台数"
    chart.height, chart.width = 11, 26
    chart.add_data(Reference(ws, min_col=2, max_col=2, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "G4")

    line = LineChart()
    line.title = "年式別 価格中央値と走行中央値"
    line.y_axis.title = "万円 / 万km"
    line.height, line.width = 11, 26
    line.add_data(Reference(ws, min_col=3, max_col=4, min_row=start, max_row=last),
                  titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(line, "G27")

    ws.column_dimensions["A"].width = 10
    for col in "BCD":
        ws.column_dimensions[col].width = 18


# ------------------------------------------------------------- 3. 車種別


def model_sheet(ws: Worksheet, listings: list[dict[str, Any]]) -> None:
    _title(
        ws,
        "在庫の多い車種の価格帯",
        "最安・中央値・最高を並べたもの。幅が広い車種は程度の差が大きい"
        "＝安い個体には理由があることが多い。狭い車種は相場が固まっている。",
    )

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in listings:
        groups[f"{row.get('maker') or ''} {row.get('model_name') or ''}".strip()].append(row)

    records = []
    for name, group in groups.items():
        prices = sorted(p for p in (_price(r) for r in group) if p)
        if len(prices) < MIN_AUCTION_SAMPLES:
            continue
        records.append((name, len(group), prices[0],
                        round(st.median(prices), 1), prices[-1]))
    records.sort(key=lambda r: -r[1])
    records = records[:TOP_MODELS]

    start = 4
    _header(ws, start, ["車種", "在庫台数", "最安(万円)", "中央値(万円)", "最高(万円)"])
    last = _rows(ws, start, records,
                 {2: INT_FMT, 3: MANYEN_FMT, 4: MANYEN_FMT, 5: MANYEN_FMT})

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = f"在庫上位{len(records)}車種の価格帯（万円）"
    chart.x_axis.title = "万円"
    chart.height, chart.width = 16, 26
    chart.add_data(Reference(ws, min_col=3, max_col=5, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "H4")

    ws.column_dimensions["A"].width = 26
    for col in "BCDE":
        ws.column_dimensions[col].width = 15


# ----------------------------------------------------------- 4. 落札比較


def auction_sheet(ws: Worksheet, listings: list[dict[str, Any]],
                  auctions: list[dict[str, Any]]) -> None:
    """同じ車種で 店頭 と ヤフオク落札 を並べる。

    名寄せはメーカー＋車種名の文字列でやる。カーセンサーとヤフオクで
    車種の切り方が違う（例: カーセンサーは「240」と「240エステート」を分けるが
    ヤフオクは「240」ひとつ）ので、突き合わせは前方一致で緩めに取り、
    サンプルが薄いものは載せない。
    """
    _title(
        ws,
        "店頭価格 と ヤフオク落札価格 の差",
        "同じ車種の 支払総額中央値（カーセンサー）と 落札中央値（ヤフオク・過去180日）。"
        "差が大きい車種ほど、オークションで買う旨みが大きい。"
        "ただし現車確認ができず、名義変更・陸送・整備は自分で手配することになる。",
    )

    shop: dict[str, list[float]] = defaultdict(list)
    for row in listings:
        price = _price(row)
        if price and row.get("model_name"):
            shop[f"{row.get('maker') or ''} {row['model_name']}".strip()].append(price)

    auction: dict[str, list[float]] = defaultdict(list)
    for row in auctions:
        if row.get("price") and row.get("model_name"):
            key = f"{row.get('maker') or ''} {row['model_name']}".strip()
            auction[key].append(row["price"] / MANYEN)

    records = []
    for name, shop_prices in shop.items():
        # ヤフオク側は車種の切り方が粗いので、前方一致で拾う
        hits = [p for k, v in auction.items()
                if k == name or name.startswith(k) or k.startswith(name)
                for p in v]
        if len(hits) < MIN_AUCTION_SAMPLES or len(shop_prices) < MIN_AUCTION_SAMPLES:
            continue
        shop_median = st.median(shop_prices)
        auction_median = st.median(hits)
        if not auction_median:
            continue
        records.append((
            name, len(shop_prices), round(shop_median, 1), len(hits),
            round(auction_median, 1), round(shop_median - auction_median, 1),
            round((shop_median / auction_median - 1) * 100, 1),
        ))
    records.sort(key=lambda r: -(r[5] or 0))
    records = records[:TOP_MODELS]

    if not records:
        ws["A4"] = "突き合わせできる車種がありませんでした（ヤフオク側のサンプル不足）。"
        return

    start = 4
    _header(ws, start, ["車種", "店頭\n台数", "店頭中央値\n(万円)", "落札\n件数",
                        "落札中央値\n(万円)", "差\n(万円)", "店頭が高い\n(%)"])
    last = _rows(ws, start, records,
                 {2: INT_FMT, 3: MANYEN_FMT, 4: INT_FMT, 5: MANYEN_FMT,
                  6: MANYEN_FMT, 7: "0.0"})

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = "店頭価格 vs ヤフオク落札価格（中央値・万円）"
    chart.x_axis.title = "万円"
    chart.height, chart.width = 16, 26
    chart.add_data(Reference(ws, min_col=3, max_col=3, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.add_data(Reference(ws, min_col=5, max_col=5, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "J4")

    ws.column_dimensions["A"].width = 26
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[start].height = 40


def build(wb, listings: list[dict[str, Any]], auctions: list[dict[str, Any]],
          *, year_from: int, year_to: int) -> None:
    maker_sheet(wb.create_sheet("グラフ_メーカー別"), listings)
    year_sheet(wb.create_sheet("グラフ_年式別"), listings, year_from, year_to)
    model_sheet(wb.create_sheet("グラフ_車種別"), listings)
    if auctions:
        auction_sheet(wb.create_sheet("グラフ_落札比較"), listings, auctions)
    log.info("  グラフシート %s 枚を追加", 4 if auctions else 3)
