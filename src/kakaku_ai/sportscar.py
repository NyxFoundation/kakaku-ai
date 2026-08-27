"""スポーツカー版の相場ブック。**落札の意思決定を支える**ことに振ってある。

他の 4 冊が「どの車種を買うか」を決めるための本なのに対して、こちらは
車種はもう決まっている人が **「この個体にいくらまで入れるか」** を決める本。
なので中心に置くのは店頭の掲載価格ではなく、**実際にいくらで落ちたか**。

### なぜヤフオクの落札が効くか

スポーツカーのヤフオクは **出品者の 93〜100% が個人**（実測）。業者の販路では
なく純粋な個人間市場で、店頭とは別の需給で動いている。しかも業者
オークション（USS / TAA）と違って会員でなくても全数が見られる。
それでいて誰も全数を数えていない。

180 日ぶんを取りこぼしなく取っているので、車種 × 年式 × 走行距離で
「いくらで落ちているか」の分布がそのまま出せる。

### 上限の出し方

年式ごとの落札中央値に、**その年式のなかでの走行距離の効き**を乗せる。
車種をまたいだ回帰はしない（240 の 8万km と V70 の 8万km が別物なのと同じで、
コペンの 5万km と RX-7 の 5万km は意味が違う）。

出すのは 25% / 中央値 / 75% の 3 本。「中央値まで」なら普通に競り負けるし、
「75% まで」出せば大抵は獲れるが高値掴みになる、という幅を見せるため。
1 本の「適正価格」を出すより、幅と件数を見せたほうが判断に使える。

### 終了タイミングについて（先に結論）

「人が見ていない時間に終わる出品は安く獲れる」という話をよく聞くので調べたが、
**この 180 日ぶんのデータでは差が出ていない。**

* 落札の **80% が 18〜24時に終わっている**。そもそも比較できるほど他の時間帯に
  玉が無い（深夜は 21件しかない）
* 曜日別の落札額指数は 95〜105 の範囲に収まり、比較の土台を厳しくすると
  順位が入れ替わる。つまりノイズ

シートにはこの否定的な結果をそのまま載せている。効くと言えないことを
効くように見せるほうが害が大きいため。
"""

from __future__ import annotations

import logging
import statistics as st
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from . import store
from .charts import BORDER, INT_FMT, MANYEN_FMT, _header, _median, _title
from .excel import TITLE_FONT, _write_table
from .vehicles import DATA_DIR
from .wide import load_catalog

log = logging.getLogger(__name__)

DEFAULT_OUTPUT = DATA_DIR / "xlsx" / "souba_sportscar.xlsx"
MANYEN = 10_000

# カーセンサーのボディタイプでは拾いきれない車種。GT-R はセダン扱い、
# ランエボ／WRX もセダン、シビックタイプR はハッチバックに入っている
SPORTS_NAMEPLATES = (
    "GT-R", "RX-7", "RX-8", "スープラ", "MR2", "MR-S", "ロードスター", "シルビア",
    "180SX", "フェアレディZ", "NSX", "S2000", "インテグラ", "タイプR",
    "ランサーエボリューション", "インプレッサ", "WRX", "GTO", "FTO", "セリカ",
    "86", "BRZ", "コペン", "カプチーノ", "ビート", "AZ-1", "アルテッツァ",
    "ロータス", "ケイマン", "ボクスター", "911", "コルベット", "バイパー",
)
SPORTS_BODIES = ("クーペ", "オープン")

# 分布を出すのに最低限ほしい件数。これ未満は「参考」扱いで別に出す
MIN_SAMPLES = 3
# 走行距離の係数はこれ未満だと符号すら当てにならない。実測で、落札 4〜7件だと
# 39% が「走るほど高い」という有り得ない符号になった（15件以上なら 10%）
MIN_MILEAGE_SAMPLES = 10


def is_sports(name: str | None, body_type: str | None) -> bool:
    name = name or ""
    return bool(name) and (
        any(tag in name for tag in SPORTS_NAMEPLATES) or body_type in SPORTS_BODIES
    )


def _body_index() -> dict[tuple[str, str], str]:
    """(メーカー, 車種) → ボディタイプ。ヤフオク側にはボディタイプが無いので補う。"""
    return {
        (m.get("maker") or "", m.get("model_name") or ""): m.get("body_type") or ""
        for m in load_catalog().values()
    }


def collect() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """(スポーツカーの落札, スポーツカーの店頭在庫) を保存済みデータから拾う。"""
    body = _body_index()
    auctions, _ = store.read_latest("yahoo_used_cars")
    sports_auctions = [
        r for r in auctions
        if r.get("price") and is_sports(r.get("model_name"),
                                        body.get((r.get("maker") or "", r.get("model_name") or "")))
    ]
    for row in sports_auctions:
        row["body_type"] = body.get((row.get("maker") or "", row.get("model_name") or ""))

    stock, _ = store.read_latest("classic_listings")
    sports_stock = [r for r in stock if is_sports(r.get("model_name"), r.get("body_type"))]
    return sports_auctions, sports_stock


# ------------------------------------------------------- 入札の目安


def _mileage_coefficient(rows: list[dict[str, Any]]) -> float | None:
    """同じ車種・同じ年式のなかで、走行 1万km あたり価格が何 % 変わるか。

    年式をまたぐと「古い＝安い＝よく走っている」が混ざって、走行距離の
    効きを過大に見積もる。年式のなかだけで見る。
    """
    pairs = [(r["mileage_km"] / MANYEN, r["price"]) for r in rows
             if r.get("mileage_km") and r.get("price")]
    if len(pairs) < MIN_MILEAGE_SAMPLES:
        return None
    xs = [p[0] for p in pairs]
    ys = [p[1] for p in pairs]
    mx, my = sum(xs) / len(xs), sum(ys) / len(ys)
    var = sum((x - mx) ** 2 for x in xs)
    if var == 0 or my == 0:
        return None
    slope = sum((x - mx) * (y - my) for x, y in pairs) / var
    return round(slope / my * 100, 1)


def bid_guide(auctions: list[dict[str, Any]],
              stock: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """車種 × 年式で「いくらまで入れるか」の目安を作る。"""
    groups: dict[tuple[str, str, int], list[dict[str, Any]]] = defaultdict(list)
    for row in auctions:
        if row.get("model_year"):
            groups[(row.get("maker") or "", row["model_name"], row["model_year"])].append(row)

    shop: dict[tuple[str, int], list[float]] = defaultdict(list)
    for row in stock:
        price = row.get("total_price_manyen") or row.get("base_price_manyen")
        if price and row.get("model_year"):
            shop[(row.get("model_name") or "", row["model_year"])].append(price)

    out: list[dict[str, Any]] = []
    for (maker, model, year), rows in groups.items():
        prices = sorted(r["price"] / MANYEN for r in rows)
        mileages = [r["mileage_km"] for r in rows if r.get("mileage_km")]
        bids = [r["bid_count"] or 0 for r in rows]
        shop_prices = shop.get((model, year)) or []
        shop_median = _median(shop_prices) if len(shop_prices) >= 2 else None
        median = _median(prices)
        out.append({
            "maker": maker,
            "model_name": model,
            "body_type": rows[0].get("body_type"),
            "model_year": year,
            "n": len(rows),
            "p25_manyen": round(prices[len(prices) // 4], 1),
            "median_manyen": round(median, 1) if median else None,
            "p75_manyen": round(prices[len(prices) * 3 // 4], 1),
            "max_manyen": round(prices[-1], 1),
            "mileage_median_km": int(st.median(mileages)) if mileages else None,
            "mileage_coef_pct": _mileage_coefficient(rows),
            "bid_median": int(st.median(bids)) if bids else None,
            "individual_pct": round(
                sum(1 for r in rows if r.get("seller_is_store") is False) / len(rows) * 100),
            "shop_n": len(shop_prices) or None,
            "shop_median_manyen": round(shop_median, 1) if shop_median else None,
            "shop_premium_pct": (round((shop_median / median - 1) * 100)
                                 if shop_median and median else None),
        })
    out.sort(key=lambda r: (r["model_name"], -(r["model_year"] or 0)))
    return out


# ------------------------------------------------------- 終了タイミング


WEEKDAYS = ("月", "火", "水", "木", "金", "土", "日")
SLOTS = ((0, 6, "深夜 0-6時"), (6, 12, "午前 6-12時"),
         (12, 18, "昼 12-18時"), (18, 24, "夜 18-24時"))


def _slot(hour: int) -> str:
    return next(label for lo, hi, label in SLOTS if lo <= hour < hi)


def timing_table(auctions: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """終了の (曜日別の落札額指数, 時間帯別の件数) を返す。

    金額をそのまま平均すると「高い車がたまたま日曜に終わった」で歪むので、
    各落札を **同じ車種・同じ年式の中央値で割った比** にしてから集計する。

    比較の土台になる中央値は 5 件以上ある年式からしか取らない。3 件だと
    中央値がその落札自身になりがちで、比が 1.0 に張り付いて差が消える。
    """
    peer: dict[tuple[str, int], list[float]] = defaultdict(list)
    for row in auctions:
        if row.get("model_year") and row.get("price"):
            peer[(row["model_name"], row["model_year"])].append(row["price"])
    medians = {k: st.median(v) for k, v in peer.items() if len(v) >= 5}

    by_weekday: dict[str, list[float]] = defaultdict(list)
    by_slot: dict[str, int] = defaultdict(int)
    for row in auctions:
        if not row.get("end_time"):
            continue
        try:
            ended = datetime.fromisoformat(row["end_time"].replace("Z", "+00:00"))
        except ValueError:
            continue
        by_slot[_slot(ended.hour)] += 1
        base = medians.get((row.get("model_name"), row.get("model_year")))  # type: ignore[arg-type]
        if base:
            by_weekday[WEEKDAYS[ended.weekday()]].append(row["price"] / base)

    weekday_rows = [
        {"weekday": w,
         "n": len(by_weekday.get(w) or []),
         "index": round(st.median(by_weekday[w]) * 100) if by_weekday.get(w) else None}
        for w in WEEKDAYS
    ]
    total = sum(by_slot.values()) or 1
    slot_rows = [
        {"slot": label, "n": by_slot.get(label, 0),
         "share_pct": round(by_slot.get(label, 0) / total * 100, 1)}
        for _, _, label in SLOTS
    ]
    return weekday_rows, slot_rows


# ------------------------------------------------------------------ シート


GUIDE_COLUMNS: list[tuple[str, str]] = [
    ("maker", "メーカー"),
    ("model_name", "車種"),
    ("body_type", "ボディ\nタイプ"),
    ("model_year", "年式"),
    ("n", "落札\n件数"),
    ("p25_manyen", "安く獲れた線\n25%(万円)"),
    ("median_manyen", "普通に競ると\n中央値(万円)"),
    ("p75_manyen", "ほぼ獲れる線\n75%(万円)"),
    ("max_manyen", "最高\n(万円)"),
    ("mileage_median_km", "落札車の\n走行中央値(km)"),
    ("mileage_coef_pct", "走行1万km\nあたり(%)"),
    ("bid_median", "入札数\n中央値"),
    ("individual_pct", "個人出品\n(%)"),
    ("shop_n", "店頭\n在庫数"),
    ("shop_median_manyen", "店頭中央値\n(万円)"),
    ("shop_premium_pct", "店頭が高い\n(%)"),
]

GUIDE_FORMATS = {
    "model_year": "0", "n": INT_FMT,
    "p25_manyen": MANYEN_FMT, "median_manyen": MANYEN_FMT,
    "p75_manyen": MANYEN_FMT, "max_manyen": MANYEN_FMT,
    "mileage_median_km": INT_FMT, "mileage_coef_pct": "+0.0;-0.0;0.0",
    "bid_median": INT_FMT, "individual_pct": INT_FMT,
    "shop_n": INT_FMT, "shop_median_manyen": MANYEN_FMT, "shop_premium_pct": INT_FMT,
}

AUCTION_COLUMNS: list[tuple[str, str]] = [
    ("end_date", "終了日"),
    ("weekday", "曜日"),
    ("end_hour", "終了\n時刻"),
    ("maker", "メーカー"),
    ("model_name", "車種"),
    ("model_year", "年式"),
    ("price_manyen", "落札額\n(万円)"),
    ("vs_peer_pct", "同年式比\n(%)"),
    ("buy_now_manyen", "即決\n(万円)"),
    ("mileage_mankm", "走行\n(万km)"),
    ("repair_label", "修復歴"),
    ("mileage_label", "メーター"),
    ("bid_count", "入札数"),
    ("seller_label", "出品者"),
    ("title", "タイトル"),
    ("url", "URL"),
]

AUCTION_FORMATS = {
    "model_year": "0", "price_manyen": MANYEN_FMT, "vs_peer_pct": INT_FMT,
    "buy_now_manyen": MANYEN_FMT, "mileage_mankm": "0.0", "bid_count": INT_FMT,
}

REPAIR_LABELS = {"NONE": "なし", "REPAIRED": "あり", "EXISTS": "あり", "UNKNOWN": "わからない"}
MILEAGE_LABELS = {"REAL_MILEAGE": "実走行", "METER_REPLACEMENT": "メーター交換",
                  "UNKNOWN_MILEAGE": "不明"}


def _flatten_auctions(auctions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    peer: dict[tuple[str, int], list[float]] = defaultdict(list)
    for row in auctions:
        if row.get("model_year") and row.get("price"):
            peer[(row["model_name"], row["model_year"])].append(row["price"])
    medians = {k: st.median(v) for k, v in peer.items() if len(v) >= MIN_SAMPLES}

    out = []
    for row in auctions:
        end = row.get("end_time") or ""
        try:
            ended = datetime.fromisoformat(end.replace("Z", "+00:00")) if end else None
        except ValueError:
            ended = None
        base = medians.get((row.get("model_name"), row.get("model_year")))  # type: ignore[arg-type]
        is_store = row.get("seller_is_store")
        out.append({
            **row,
            "end_date": end[:10],
            "weekday": WEEKDAYS[ended.weekday()] if ended else "",
            "end_hour": f"{ended.hour:02d}:{ended.minute:02d}" if ended else "",
            "price_manyen": round(row["price"] / MANYEN, 1),
            "vs_peer_pct": round(row["price"] / base * 100) if base else None,
            "buy_now_manyen": (round(row["buy_now_price"] / MANYEN, 1)
                               if row.get("buy_now_price") else None),
            "mileage_mankm": (round(row["mileage_km"] / MANYEN, 1)
                              if row.get("mileage_km") else None),
            "repair_label": REPAIR_LABELS.get(row.get("repair_type") or "", "不明"),
            "mileage_label": MILEAGE_LABELS.get(row.get("mileage_type") or "", ""),
            "seller_label": "不明" if is_store is None else ("業者" if is_store else "個人"),
        })
    out.sort(key=lambda r: r.get("end_date") or "", reverse=True)
    return out


def _readme(ws, auctions, stock, guide, snapshot) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 108
    ws["A1"] = "スポーツカー 落札相場（入札の判断材料）"
    ws["A1"].font = TITLE_FONT

    models = len({r["model_name"] for r in auctions})
    individual = sum(1 for r in auctions if r.get("seller_is_store") is False)
    lines = [
        ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("取得時点", snapshot or "不明"),
        ("収録", f"ヤフオク落札 {len(auctions)}台 / {models}車種（過去180日・全数）"),
        ("出品者の内訳", f"個人 {individual}台（{individual / max(len(auctions), 1) * 100:.0f}%）"
                    f" / 業者ほか {len(auctions) - individual}台"),
        ("店頭在庫", f"{len(stock)}台（カーセンサー・1988〜2001年式）"),
        ("", ""),
        ("■ この本の使い方", ""),
        ("入札の目安",
         "車種 × 年式で「いくらで落ちているか」を 25% / 中央値 / 75% の3本で出したもの。"
         "1本の適正価格は出していない。中央値までしか出さなければ普通に競り負けるし、"
         "75%まで出せば大抵は獲れるが高値掴みになる。その幅と件数を見て決めるための表。"),
        ("落札明細",
         "1台ずつの落札結果。同年式比(%) が付いているので、相場より安く落ちた個体を"
         "並べ替えて眺められる。安く落ちた出品に共通点（終了時刻・写真枚数・"
         "説明の書き方）があれば、それが狙い目のパターンになる。"),
        ("グラフ_終了タイミング",
         "「終了時刻を狙えば安く獲れる」の検証。結果は**差が出ていない**。"
         "落札の80%が夜18-24時に集中していて他の時間帯に玉が無く、曜日別の"
         "指数も95〜105のノイズ範囲。否定的な結果もそのまま載せている。"),
        ("グラフ_車種別", "車種ごとの落札相場と入札の厚み。"),
        ("グラフ_年式別", "年式ごとの落札中央値。値落ちが止まる年式＝底が見える。"),
        ("", ""),
        ("■ 数字の作り方と限界", ""),
        ("落札データ",
         "ヤフオク「中古車・新車」ノードの落札を180日ぶん全数。30,382件を取りこぼし"
         "なく取得したうえで、車種名とボディタイプでスポーツカーを抜いている。"
         "ページングが15,050件で頭打ちになるので、並び順を変えて複数回さらって名寄せした。"),
        ("180日という制限",
         "ヤフオクの落札検索は終了180日ぶんしか返さない。**過去には遡れない。**"
         "週次で撮り続けて auction_id で名寄せすれば実効期間は伸びていくが、"
         "いま手元にあるのは180日ぶんが上限。季節性を語るには1年以上の蓄積が要る。"),
        ("走行距離の効き",
         "同じ車種・同じ年式のなかだけで見ている。年式をまたぐと"
         "「古い＝安い＝よく走っている」が混ざって効きを過大に見積もるため。"
         "**10件未満の年式は係数を出していない。**落札4〜7件だと39%が"
         "「走るほど高い」という有り得ない符号になった（15件以上なら10%）。"
         "符号が正の行が残っていたらサンプル不足かグレード構成の偏りを疑うこと。"),
        ("店頭との比較",
         "店頭在庫はいまのところ1988〜2001年式ぶんしか個体単位で取っていない。"
         "2002年以降の店頭中央値が空なのはそのため。落札側は全年式そろっている。"),
        ("修復歴",
         "落札明細には修復歴ありも含めて全部載せている。相場を作るときは"
         "除外すべきだが、「修復歴ありがいくらで落ちるか」自体が判断材料になるため。"),
        ("入札数",
         "終了後の総数しか取れていない。「いつ・いくらで入ったか」の過程は"
         "出品中を定期的に撮らないと分からない（入札履歴ページは robots.txt で禁止）。"
         "そこは今後の課題。"),
    ]
    for i, (key, value) in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=key)
        if key.startswith("■"):
            cell.font = TITLE_FONT
        body = ws.cell(row=i, column=2, value=value)
        body.alignment = body.alignment.copy(wrap_text=True, vertical="top")


def _rows(ws, start: int, records: list[tuple], formats: dict[int, str]) -> int:
    for offset, record in enumerate(records):
        for col_index, value in enumerate(record, start=1):
            cell = ws.cell(row=start + 1 + offset, column=col_index, value=value)
            cell.border = BORDER
            if col_index in formats:
                cell.number_format = formats[col_index]
    return start + len(records)


def _model_chart(ws, auctions: list[dict[str, Any]]) -> None:
    _title(ws, "車種別 落札相場と入札の厚み",
           "入札数はその車種に何人が張り付いているかの代理指標。"
           "同じ価格帯でも入札が薄い車種は、条件次第で安く獲れる。")

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in auctions:
        groups[f"{row.get('maker') or ''} {row['model_name']}".strip()].append(row)

    records = []
    for name, rows in groups.items():
        if len(rows) < MIN_SAMPLES:
            continue
        prices = sorted(r["price"] / MANYEN for r in rows)
        bids = [r["bid_count"] or 0 for r in rows]
        records.append((name, len(rows), round(prices[len(prices) // 4], 1),
                        round(st.median(prices), 1), round(prices[len(prices) * 3 // 4], 1),
                        int(st.median(bids))))
    records.sort(key=lambda r: -r[1])
    records = records[:25]

    start = 4
    _header(ws, start, ["車種", "落札件数", "25%(万円)", "中央値(万円)", "75%(万円)", "入札数中央値"])
    last = _rows(ws, start, records,
                 {2: INT_FMT, 3: MANYEN_FMT, 4: MANYEN_FMT, 5: MANYEN_FMT, 6: INT_FMT})

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = f"車種別 落札価格帯（上位{len(records)}車種・万円）"
    chart.x_axis.title = "万円"
    chart.height, chart.width = 18, 26
    chart.add_data(Reference(ws, min_col=3, max_col=5, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "I4")

    bids_chart = BarChart()
    bids_chart.type = "bar"
    bids_chart.title = "入札数の中央値"
    bids_chart.x_axis.title = "入札数"
    bids_chart.height, bids_chart.width = 18, 14
    bids_chart.add_data(Reference(ws, min_col=6, max_col=6, min_row=start, max_row=last),
                        titles_from_data=True)
    bids_chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(bids_chart, "U4")

    ws.column_dimensions["A"].width = 26
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15


def _year_chart(ws, auctions: list[dict[str, Any]]) -> None:
    _title(ws, "年式別 落札中央値（掲載の多い車種）",
           "値落ちが止まって横ばい、あるいは上がりはじめる年式が「底」。"
           "そこが買い時で、そこより古いものは趣味性で値が戻っている。")

    counts: dict[str, int] = defaultdict(int)
    for row in auctions:
        counts[row["model_name"]] += 1
    models = [m for m, _ in sorted(counts.items(), key=lambda kv: -kv[1])[:12]]

    table: dict[str, dict[int, float]] = defaultdict(dict)
    years: set[int] = set()
    for row in auctions:
        if row["model_name"] in models and row.get("model_year"):
            table[row["model_name"]].setdefault(row["model_year"], [])  # type: ignore[arg-type]
    grouped: dict[tuple[str, int], list[float]] = defaultdict(list)
    for row in auctions:
        if row["model_name"] in models and row.get("model_year"):
            grouped[(row["model_name"], row["model_year"])].append(row["price"] / MANYEN)
            years.add(row["model_year"])

    year_list = sorted(y for y in years if y >= 1988)
    start = 4
    _header(ws, start, ["車種", *[str(y) for y in year_list]])
    written = 0
    for model in models:
        cells = {y: grouped.get((model, y)) for y in year_list}
        if not any(v and len(v) >= 2 for v in cells.values()):
            continue
        written += 1
        ws.cell(row=start + written, column=1, value=model).border = BORDER
        for col, year in enumerate(year_list, start=2):
            values = cells.get(year) or []
            cell = ws.cell(row=start + written, column=col,
                           value=round(st.median(values), 1) if len(values) >= 2 else None)
            cell.number_format = MANYEN_FMT
            cell.border = BORDER

    if written:
        ws.conditional_formatting.add(
            f"B{start + 1}:{get_column_letter(len(year_list) + 1)}{start + written}",
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           mid_type="percentile", mid_value=50, mid_color="BDD7EE",
                           end_type="max", end_color="2E75B6"),
        )
    ws.freeze_panes = f"B{start + 1}"
    ws.column_dimensions["A"].width = 24
    for col in range(2, len(year_list) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 8


def _timing_chart(ws, auctions: list[dict[str, Any]]) -> None:
    weekday_rows, slot_rows = timing_table(auctions)
    _title(
        ws,
        "終了タイミングと落札額 — 差は出ていない",
        "「人が見ていない時間に終わる出品は安い」を検証した結果。"
        "落札の80%が夜18-24時に集中していて、他の時間帯はそもそも玉が無い。"
        "曜日別の指数（同じ車種・同じ年式の中央値＝100）も95〜105に収まり、"
        "比較の土台を厳しくすると順位が入れ替わる。つまりノイズで、"
        "終了タイミングを狙う戦術の根拠にはならない。",
    )

    start = 4
    _header(ws, start, ["終了時間帯", "落札件数", "構成比(%)"])
    last = _rows(ws, start,
                 [(r["slot"], r["n"], r["share_pct"]) for r in slot_rows],
                 {2: INT_FMT, 3: "0.0"})

    chart = BarChart()
    chart.type = "col"
    chart.title = "終了時間帯の分布（夜に集中している）"
    chart.y_axis.title = "落札件数"
    chart.height, chart.width = 10, 20
    chart.add_data(Reference(ws, min_col=2, max_col=2, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "F4")

    start2 = last + 3
    ws.cell(row=start2 - 1, column=1,
            value="曜日別 落札額指数（同じ車種・同じ年式の中央値＝100）").font = TITLE_FONT
    _header(ws, start2, ["終了曜日", "件数", "指数"])
    last2 = _rows(ws, start2,
                  [(r["weekday"], r["n"], r["index"]) for r in weekday_rows],
                  {2: INT_FMT, 3: INT_FMT})

    line = LineChart()
    line.title = "曜日別 落札額指数（100 = 同年式の中央値どおり）"
    line.y_axis.title = "指数"
    line.height, line.width = 10, 20
    line.add_data(Reference(ws, min_col=3, max_col=3, min_row=start2, max_row=last2),
                  titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=start2 + 1, max_row=last2))
    ws.add_chart(line, "F24")

    ws.column_dimensions["A"].width = 20
    for col in "BC":
        ws.column_dimensions[col].width = 14


def build(output: Path = DEFAULT_OUTPUT) -> Path:
    auctions, stock = collect()
    if not auctions:
        raise RuntimeError("ヤフオク落札データがありません。先に落札の全数取得を実行してください。")
    _, snapshot = store.read_latest("yahoo_used_cars")
    guide = bid_guide(auctions, stock)

    wb = Workbook()
    _readme(wb.active, auctions, stock, guide, snapshot)
    wb.active.title = "README"

    _model_chart(wb.create_sheet("グラフ_車種別"), auctions)
    _year_chart(wb.create_sheet("グラフ_年式別"), auctions)
    _timing_chart(wb.create_sheet("グラフ_終了タイミング"), auctions)

    # 件数の少ない年式は判断に使えないので下に沈める
    _write_table(wb.create_sheet("入札の目安"), GUIDE_COLUMNS,
                 sorted(guide, key=lambda r: (-(r["n"] >= MIN_SAMPLES), r["model_name"],
                                              -(r["model_year"] or 0))),
                 number_formats=GUIDE_FORMATS)
    _write_table(wb.create_sheet("落札明細"), AUCTION_COLUMNS,
                 _flatten_auctions(auctions), number_formats=AUCTION_FORMATS,
                 wrap_columns={"title"})

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    log.info("%s（落札 %s台 / %s車種 / 目安 %s行）",
             output.name, len(auctions), len({r["model_name"] for r in auctions}), len(guide))
    return output
