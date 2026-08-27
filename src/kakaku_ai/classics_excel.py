"""旧車在庫の xlsx。本体の相場ブックとは別ファイルにする。

`souba_minivan.xlsx` は「車種 × 年式の相場」を積み上げる時系列のブックで、
こちらは「いま買える 1 台」を並べる在庫のブック。粒度も使い方も違うので
混ぜない。フィルタで メーカー / 国産・輸入 / ボディタイプ / 車種 / 年式 /
価格 / 走行距離 / 修復歴 を絞れるようにしてある。
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from . import classics_charts
from .excel import ALT_FILL, INT_FMT, MANYEN_FMT, TITLE_FONT, _write_table
from .vehicles import DATA_DIR

log = logging.getLogger(__name__)

DEFAULT_OUTPUT = DATA_DIR / "xlsx" / "souba_classics.xlsx"

LISTING_COLUMNS: list[tuple[str, str]] = [
    ("maker", "メーカー"),
    ("origin", "国産/輸入"),
    ("body_type", "ボディタイプ\n(用途)"),
    ("model_name", "車種"),
    ("model_year", "年式"),
    ("score", "状態スコア"),
    ("total_price_manyen", "支払総額\n(万円)"),
    ("base_price_manyen", "車両本体\n(万円)"),
    ("mileage_mankm", "走行距離\n(万km)"),
    ("repair_history", "修復歴"),
    ("inspection", "車検"),
    ("inspection_left_months", "車検残\n(ヶ月)"),
    ("warranty", "保証"),
    ("why", "評価の理由"),
    ("title", "グレード・装備"),
    ("peer_count", "同年式の\n在庫数"),
    ("production_period", "生産期間"),
    ("url", "掲載URL"),
]

LISTING_FORMATS = {
    "model_year": "0",
    "score": "+0.0;-0.0;0.0",
    "total_price_manyen": MANYEN_FMT,
    "base_price_manyen": MANYEN_FMT,
    "mileage_mankm": "0.0",
    "inspection_left_months": INT_FMT,
    "peer_count": INT_FMT,
}

MODEL_COLUMNS: list[tuple[str, str]] = [
    ("maker", "メーカー"),
    ("origin", "国産/輸入"),
    ("body_type", "ボディタイプ\n(用途)"),
    ("model_name", "車種"),
    ("listing_count", "在庫台数"),
    ("year_min", "最古\n年式"),
    ("year_max", "最新\n年式"),
    ("price_min_manyen", "最安\n(万円)"),
    ("price_median_manyen", "中央値\n(万円)"),
    ("price_max_manyen", "最高\n(万円)"),
    ("price_unknown_n", "価格応談\n(台)"),
    ("mileage_median_mankm", "走行中央値\n(万km)"),
    ("repaired_n", "修復歴あり\n(台)"),
    ("best_score", "最高\nスコア"),
    ("production_period", "生産期間"),
    ("carsensor_code", "コード"),
    ("url", "一覧URL"),
]

AUCTION_COLUMNS: list[tuple[str, str]] = [
    ("end_date", "終了日"),
    ("maker", "メーカー"),
    ("model_name", "車種"),
    ("model_year", "年式"),
    ("price_manyen", "落札額\n(万円)"),
    ("buy_now_manyen", "即決\n(万円)"),
    ("mileage_mankm", "走行距離\n(万km)"),
    ("repair_label", "修復歴"),
    ("mileage_label", "メーター"),
    ("bid_count", "入札数"),
    ("seller_label", "出品者"),
    ("title", "タイトル"),
    ("url", "落札URL"),
]

AUCTION_FORMATS = {
    "model_year": "0",
    "price_manyen": MANYEN_FMT,
    "buy_now_manyen": MANYEN_FMT,
    "mileage_mankm": "0.0",
    "bid_count": INT_FMT,
}

REPAIR_LABELS = {"NONE": "なし", "REPAIRED": "あり", "EXISTS": "あり", "UNKNOWN": "わからない"}
MILEAGE_LABELS = {"REAL_MILEAGE": "実走行", "METER_REPLACEMENT": "メーター交換",
                  "UNKNOWN_MILEAGE": "不明"}

MODEL_FORMATS = {
    "listing_count": INT_FMT,
    "year_min": "0",
    "year_max": "0",
    "price_min_manyen": MANYEN_FMT,
    "price_median_manyen": MANYEN_FMT,
    "price_max_manyen": MANYEN_FMT,
    "price_unknown_n": INT_FMT,
    "mileage_median_mankm": "0.0",
    "repaired_n": INT_FMT,
    "best_score": "+0.0;-0.0;0.0",
}


def _flatten(row: dict[str, Any]) -> dict[str, Any]:
    """シートに出す形に整える。生の km と理由リストはそのままでは読めない。"""
    from .aggregate import months_until

    out = dict(row)
    out["mileage_mankm"] = round(row["mileage_km"] / 10000, 1) if row.get("mileage_km") else None
    out["inspection_left_months"] = months_until(row.get("inspection_ym") or 0)
    out["why"] = " / ".join(row.get("why") or [])
    out["score"] = round(row.get("score") or 0, 1)
    return out


def _flatten_auction(row: dict[str, Any]) -> dict[str, Any]:
    out = dict(row)
    out["end_date"] = (row.get("end_time") or "")[:10]
    out["price_manyen"] = round(row["price"] / 10000, 1) if row.get("price") else None
    out["buy_now_manyen"] = (round(row["buy_now_price"] / 10000, 1)
                             if row.get("buy_now_price") else None)
    out["mileage_mankm"] = round(row["mileage_km"] / 10000, 1) if row.get("mileage_km") else None
    out["repair_label"] = REPAIR_LABELS.get(row.get("repair_type") or "", "不明")
    out["mileage_label"] = MILEAGE_LABELS.get(row.get("mileage_type") or "", "")
    # 個人出品かどうかは旧車ではとくに効く（業者は整備済み、個人は現状渡しが多い）
    is_store = row.get("seller_is_store")
    out["seller_label"] = "不明" if is_store is None else ("業者" if is_store else "個人")
    return out


def _readme(ws, listings: list[dict[str, Any]], models: list[dict[str, Any]],
            auctions: list[dict[str, Any]],
            year_from: int, year_to: int, snapshot: str) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 108
    ws["A1"] = f"旧車在庫データベース（{year_from}〜{year_to}年式・全メーカー）"
    ws["A1"].font = TITLE_FONT

    makers = len({r.get("maker") for r in models})
    imported = sum(r["listing_count"] for r in models if r.get("origin") == "輸入")
    lines = [
        ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("取得時点", snapshot),
        ("対象", f"{year_from}〜{year_to}年式 / 全国 / カーセンサー掲載在庫"),
        ("収録", f"{len(listings)}台 / {len(models)}車種 / {makers}メーカー"
                 f"（うち輸入車 {imported}台）"),
        ("ヤフオク落札", f"{len(auctions)}台（過去180日）" if auctions else "未取得"),
        ("", ""),
        ("■ シート", ""),
        ("グラフ_メーカー別", "メーカーごとの在庫台数と価格中央値。どこに玉が残っているか。"),
        ("グラフ_年式別", "年式ごとの台数・価格・走行距離。古いほど台数が減り価格は上がる。"),
        ("グラフ_車種別", "在庫の多い車種の価格帯（最安・中央値・最高）。幅が広い＝程度の差が大きい。"),
        ("グラフ_落札比較", "同じ車種の店頭価格とヤフオク落札価格の差。差が大きいほど"
                       "オークションの旨みが大きい。"),
        ("候補一覧", "カーセンサー在庫の1台ずつの明細。状態スコア順。メーカー・ボディタイプ・"
                 "年式・価格・修復歴でフィルタできる。"),
        ("ヤフオク落札", "過去180日の落札結果。実際にいくらで売れたかはこちら。"
                    "出品者が個人か業者かも出している。"),
        ("車種別サマリ", "車種ごとの在庫台数・価格帯・走行中央値。どの車種が現実的か俯瞰する用。"),
        ("メーカー別サマリ", "メーカー単位の台数と価格帯。"),
        ("", ""),
        ("■ 状態スコアの内訳", ""),
        ("走行距離",
         "同じ車種・同じ年式のなかで、中央値の50%以下なら +3.0 / 80%以下なら +1.5 / "
         "150%以上なら −1.5。車種をまたいだ比較はしない（240の8万kmとV70の8万kmは"
         "意味が違うため）。同年式に在庫が1台しかない車種は比較相手がいないので加点なし。"),
        ("修復歴", "なし +2.0 / あり −4.0"),
        ("車検", "残り月数/12（最大 +2.0）。残がなくても「車検整備付」なら +1.0"),
        ("保証", "保証付きなら +1.0"),
        ("価格",
         "同年式中央値の 60〜90% なら +1.0。50%未満は **−1.0**。"
         "30年落ちで極端に安い個体は、それなりの理由があることが多いため。"),
        ("デザイン", "数値化できないので採点していない。「グレード・装備」列と掲載URLを見て判断する。"),
        ("", ""),
        ("■ 見るときの注意", ""),
        ("価格応談",
         "支払総額を出していない店がある。低走行の人気車に多く、"
         "問い合わせると高く出る傾向。価格列が空の行がそれ。"),
        ("タイミングベルト",
         "この年代でいちばん重要。切れるとエンジンが壊れる。"
         "「グレード・装備」列に交換済みの記載がない個体は 10〜15万の上乗せを見ておく。"),
        ("車検残",
         "残がないと取り直しで 15〜25万かかる。車検残(ヶ月) 列で確認する。"),
        ("オークション相場との差",
         "同じ車がヤフオクなら店頭の 1/2〜1/3 で出ることがある。"
         "ただし現車確認ができず、名義変更・陸送・整備を自分で手配する必要がある。"),
    ]
    for i, (key, value) in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=key).font = TITLE_FONT if key.startswith("■") else None
        cell = ws.cell(row=i, column=2, value=value)
        cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")


def _maker_summary(models: list[dict[str, Any]]) -> list[dict[str, Any]]:
    import statistics as st

    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in models:
        groups.setdefault((row.get("maker") or "", row.get("origin") or ""), []).append(row)

    out = []
    for (maker, origin), group in groups.items():
        medians = [r["price_median_manyen"] for r in group if r.get("price_median_manyen")]
        out.append({
            "maker": maker,
            "origin": origin,
            "model_count": len(group),
            "listing_count": sum(r["listing_count"] for r in group),
            "price_median_manyen": round(st.median(medians), 1) if medians else None,
            "price_min_manyen": min((r["price_min_manyen"] for r in group
                                     if r.get("price_min_manyen")), default=None),
            "price_max_manyen": max((r["price_max_manyen"] for r in group
                                     if r.get("price_max_manyen")), default=None),
            "best_score": round(max(r["best_score"] for r in group), 1),
        })
    out.sort(key=lambda r: -r["listing_count"])
    return out


def build(
    listings: list[dict[str, Any]],
    models: list[dict[str, Any]],
    *,
    auctions: list[dict[str, Any]] | None = None,
    output: Path = DEFAULT_OUTPUT,
    year_from: int,
    year_to: int,
    snapshot: str,
) -> Path:
    auctions = auctions or []
    wb = Workbook()
    _readme(wb.active, listings, models, auctions, year_from, year_to, snapshot)
    wb.active.title = "README"

    # グラフは表より先に置く。開いてすぐ絵が見えるようにするため
    classics_charts.build(wb, listings, auctions, year_from=year_from, year_to=year_to)

    rows = sorted((_flatten(r) for r in listings), key=lambda r: -(r.get("score") or 0))
    ws = wb.create_sheet("候補一覧")
    n = _write_table(ws, LISTING_COLUMNS, rows,
                     number_formats=LISTING_FORMATS, wrap_columns={"why", "title"})
    if n:
        # スコア列を色分けして、良い個体が縦に見えるようにする
        col = get_column_letter(next(i for i, (k, _) in enumerate(LISTING_COLUMNS, 1) if k == "score"))
        ws.conditional_formatting.add(
            f"{col}2:{col}{n + 1}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"),
        )

    if auctions:
        ws = wb.create_sheet("ヤフオク落札")
        _write_table(
            ws, AUCTION_COLUMNS,
            sorted((_flatten_auction(r) for r in auctions),
                   key=lambda r: r.get("end_date") or "", reverse=True),
            number_formats=AUCTION_FORMATS, wrap_columns={"title"},
        )

    ws = wb.create_sheet("車種別サマリ")
    _write_table(ws, MODEL_COLUMNS, models, number_formats=MODEL_FORMATS)

    ws = wb.create_sheet("メーカー別サマリ")
    _write_table(ws, [
        ("maker", "メーカー"),
        ("origin", "国産/輸入"),
        ("model_count", "車種数"),
        ("listing_count", "在庫台数"),
        ("price_min_manyen", "最安\n(万円)"),
        ("price_median_manyen", "車種中央値の\n中央値(万円)"),
        ("price_max_manyen", "最高\n(万円)"),
        ("best_score", "最高\nスコア"),
    ], _maker_summary(models), number_formats={
        "model_count": INT_FMT, "listing_count": INT_FMT,
        "price_min_manyen": MANYEN_FMT, "price_median_manyen": MANYEN_FMT,
        "price_max_manyen": MANYEN_FMT, "best_score": "+0.0;-0.0;0.0",
    })

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    log.info("xlsx を書き出しました: %s（在庫 %s台 / %s車種 / 落札 %s台）",
             output, len(listings), len(models), len(auctions))
    return output


__all__ = ["build", "DEFAULT_OUTPUT", "ALT_FILL"]
