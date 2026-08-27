"""xlsx を 4 冊に分ける。

1 冊に全部入れると、ミニバンを選びたい人が旧車の 1 万行をスクロールすることに
なるし、シートも 20 枚を超えて開いた瞬間どこを見ればいいか分からなくなる。
用途で分ける:

| ファイル | 中身 | 誰が見るか |
|---|---|---|
| `souba_all.xlsx`      | 全 2,237 車種のカタログと年式別相場 | 車種を決めていない人 |
| `souba_minivan.xlsx`  | ミニバン。深掘り20車種は口コミ・不具合まで | ミニバンから選ぶ人 |
| `souba_standard.xlsx` | 乗用車（ミニバン・トラックを除く） | 普通の車から選ぶ人 |
| `souba_classics.xlsx` | 1988〜2001年式の旧車 | 旧車から選ぶ人 |

特化ブック（下 3 つ）は「その中から 1 台選ぶ」ための構成にする。共通で

* **車種比較** … 1 行 1 車種。価格帯・掲載台数・走行距離・値落ち率を横並びにして、
  どの車種が候補に残るかをここだけで決められるようにする
* **年式別相場** … 車種 × 年式のピボット。車種を決めたあと「どの年式が狙い目か」
* **グラフ** … 上の 2 つを絵にしたもの

を置く。`souba_all.xlsx` だけは選ぶための本ではなく索引なので、
メーカーとボディタイプの俯瞰を厚くしてある。
"""

from __future__ import annotations

import logging
import statistics as st
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from . import store
from .charts import BORDER, INT_FMT, MANYEN_FMT, _header, _title
from .excel import TITLE_FONT, _write_table
from .vehicles import DATA_DIR

log = logging.getLogger(__name__)

OUTPUT_DIR = DATA_DIR / "xlsx"

# 乗用車として扱うボディタイプ。トラック・ピックアップ・その他は
# 「車を選ぶ」文脈では別物なので外す
PASSENGER_BODIES = (
    "ハッチバック", "セダン", "クロカン・ＳＵＶ", "クーペ",
    "ステーションワゴン", "オープン", "ミニバン",
)
MINIVAN_BODIES = ("ミニバン",)
STANDARD_BODIES = tuple(b for b in PASSENGER_BODIES if b not in MINIVAN_BODIES)

TOP_MAKERS = 20
TOP_MODELS = 25
MIN_PIVOT_LISTINGS = 3  # これ未満の年式はノイズなのでピボットに出さない


# --------------------------------------------------------------- 車種比較

COMPARE_COLUMNS: list[tuple[str, str]] = [
    ("maker", "メーカー"),
    ("origin", "国産/輸入"),
    ("body_type", "ボディタイプ\n(用途)"),
    ("model_name", "車種"),
    ("production_period", "生産期間"),
    ("listing_count", "掲載台数"),
    ("shop_count", "取扱\n店舗数"),
    ("year_min", "最古\n年式"),
    ("year_max", "最新\n年式"),
    ("price_min_manyen", "最安\n(万円)"),
    ("price_median_manyen", "中央値\n(万円)"),
    ("price_max_manyen", "最高\n(万円)"),
    ("price_spread_pct", "価格の幅\n(最高/最安 倍)"),
    ("mileage_median_km", "走行中央値\n(km)"),
    ("depreciation_pct", "値落ち率\n(%/年)"),
    ("url", "相場ページ"),
]

COMPARE_FORMATS = {
    "listing_count": INT_FMT,
    "shop_count": INT_FMT,
    "year_min": "0",
    "year_max": "0",
    "price_min_manyen": MANYEN_FMT,
    "price_median_manyen": MANYEN_FMT,
    "price_max_manyen": MANYEN_FMT,
    "price_spread_pct": "0.0",
    "mileage_median_km": INT_FMT,
    "depreciation_pct": "0.0",
}


def _depreciation(rows: list[dict[str, Any]]) -> float | None:
    """新しい年式と古い年式の中央値から、1 年あたり何 % 落ちるかを出す。

    「それ以前」バケット（`is_open_bucket`）は年式が確定しないので使わない。
    2 年式ぶん無いと傾きが出せないので、その場合は None。
    """
    points = [
        (r["model_year"], r["retail_median_manyen"])
        for r in rows
        if not r.get("is_open_bucket") and r.get("retail_median_manyen")
    ]
    if len(points) < 2:
        return None
    points.sort()
    (old_year, old_price), (new_year, new_price) = points[0], points[-1]
    span = new_year - old_year
    if span <= 0 or not new_price:
        return None
    return round((1 - old_price / new_price) / span * 100, 1)


def compare_rows(
    summaries: list[dict[str, Any]],
    by_year: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """1 行 1 車種の比較表を組む。"""
    years: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in by_year:
        years[row["carsensor_code"]].append(row)

    out: list[dict[str, Any]] = []
    for summary in summaries:
        code = summary["carsensor_code"]
        rows = years.get(code, [])
        prices = sorted(r["retail_median_manyen"] for r in rows if r.get("retail_median_manyen"))
        # 「それ以前」バケットは年式が確定しないので、年式レンジには入れない
        dated = [r["model_year"] for r in rows if not r.get("is_open_bucket")]
        out.append({
            **{k: summary.get(k) for k in
               ("maker", "origin", "body_type", "model_name", "production_period",
                "listing_count", "shop_count", "url", "carsensor_code")},
            "year_min": min(dated) if dated else None,
            "year_max": max(dated) if dated else None,
            "price_min_manyen": prices[0] if prices else None,
            "price_median_manyen": round(st.median(prices), 1) if prices else None,
            "price_max_manyen": prices[-1] if prices else None,
            "price_spread_pct": (round(prices[-1] / prices[0], 1)
                                 if prices and prices[0] else None),
            "mileage_median_km": summary.get("retail_median_mileage_km"),
            "depreciation_pct": _depreciation(rows),
        })
    out.sort(key=lambda r: -(r.get("listing_count") or 0))
    return out


# --------------------------------------------------------------- 年式ピボット


def year_pivot(ws, compare: list[dict[str, Any]], by_year: list[dict[str, Any]],
               *, limit: int = 60) -> None:
    """行=車種 / 列=年式 の中央値マトリクス。行内で色を付ける。

    横に並べると「同じ車種のなかでどの年式が急に安くなるか」が一目で分かる。
    値落ちは毎年一定ではなく、モデルチェンジの前後で段差になることが多い。
    """
    _title(
        ws,
        "車種 × 年式 の小売相場（中央値・万円）",
        "掲載台数の多い順に上位%s車種。行のなかで安い年式ほど白く、高いほど濃い。"
        "段差が出ているところがモデルチェンジの境目で、その手前が狙い目になりやすい。"
        "掲載が%s台未満の年式と、年式が確定しない「それ以前／それ以降」は出していない。"
        % (limit, MIN_PIVOT_LISTINGS),
    )

    codes = [r["carsensor_code"] for r in compare[:limit]]
    names = {r["carsensor_code"]: f"{r.get('maker') or ''} {r.get('model_name') or ''}".strip()
             for r in compare[:limit]}
    wanted = set(codes)

    table: dict[str, dict[int, float]] = defaultdict(dict)
    years: set[int] = set()
    for row in by_year:
        # 「それ以前 / それ以降」バケットは年式が確定していない。しかも
        # 発売前の年に 1 台だけ載っていることがあり（スペーシアの 2012年に
        # 220万が 1 台）、そのまま出すと色スケールがその 1 台に引っぱられる。
        # ここは「どの年式が狙い目か」を見る表なので、年式が確定した行だけ使う
        if (row["carsensor_code"] in wanted
                and row.get("retail_median_manyen")
                and not row.get("is_open_bucket")
                and (row.get("listing_count") or 0) >= MIN_PIVOT_LISTINGS):
            table[row["carsensor_code"]][row["model_year"]] = row["retail_median_manyen"]
            years.add(row["model_year"])

    year_list = sorted(years)
    if not year_list:
        ws["A4"] = "年式別の相場がありません。"
        return

    start = 4
    _header(ws, start, ["車種", *[str(y) for y in year_list]])
    written = 0
    for code in codes:
        cells = table.get(code)
        if not cells:
            continue
        written += 1
        ws.cell(row=start + written, column=1, value=names[code]).border = BORDER
        for col, year in enumerate(year_list, start=2):
            cell = ws.cell(row=start + written, column=col, value=cells.get(year))
            cell.number_format = MANYEN_FMT
            cell.border = BORDER

    if written:
        ws.conditional_formatting.add(
            f"B{start + 1}:{get_column_letter(len(year_list) + 1)}{start + written}",
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           mid_type="percentile", mid_value=50, mid_color="BDD7EE",
                           end_type="max", end_color="2E75B6"),
        )
    ws.freeze_panes = f"B{start + 1}"
    ws.column_dimensions["A"].width = 26
    for col in range(2, len(year_list) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 9


# ------------------------------------------------------------------ グラフ


def _rows(ws, start: int, records: list[tuple], formats: dict[int, str]) -> int:
    for offset, record in enumerate(records):
        for col_index, value in enumerate(record, start=1):
            cell = ws.cell(row=start + 1 + offset, column=col_index, value=value)
            cell.border = BORDER
            if col_index in formats:
                cell.number_format = formats[col_index]
    return start + len(records)


def maker_chart(ws, compare: list[dict[str, Any]]) -> None:
    _title(ws, "メーカー別 車種数・掲載台数・価格",
           "掲載台数はいま買える玉の数。台数が少ないメーカーは、選ぼうにも選択肢が無い。")

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in compare:
        groups[row.get("maker") or "不明"].append(row)

    records = []
    for maker, group in groups.items():
        medians = [r["price_median_manyen"] for r in group if r.get("price_median_manyen")]
        records.append((maker, len(group), sum(r.get("listing_count") or 0 for r in group),
                        round(st.median(medians), 1) if medians else None))
    records.sort(key=lambda r: -r[2])
    records = records[:TOP_MAKERS]

    start = 4
    _header(ws, start, ["メーカー", "車種数", "掲載台数", "価格中央値(万円)"])
    last = _rows(ws, start, records, {2: INT_FMT, 3: INT_FMT, 4: MANYEN_FMT})

    chart = BarChart()
    chart.type = "bar"
    chart.title = f"メーカー別 掲載台数（上位{len(records)}）"
    chart.x_axis.title = "台数"
    chart.height, chart.width = 15, 20
    chart.add_data(Reference(ws, min_col=3, max_col=3, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "G4")

    price = BarChart()
    price.type = "bar"
    price.title = "メーカー別 価格中央値（万円）"
    price.x_axis.title = "万円"
    price.height, price.width = 15, 20
    price.add_data(Reference(ws, min_col=4, max_col=4, min_row=start, max_row=last),
                   titles_from_data=True)
    price.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(price, "S4")

    ws.column_dimensions["A"].width = 20
    for col in "BCD":
        ws.column_dimensions[col].width = 15


def body_chart(ws, compare: list[dict[str, Any]]) -> None:
    _title(ws, "ボディタイプ（用途）別",
           "同じ予算でもボディタイプで選べる車がまるで違う。まずここで用途を決める。")

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in compare:
        groups[row.get("body_type") or "不明"].append(row)

    records = []
    for body, group in groups.items():
        medians = [r["price_median_manyen"] for r in group if r.get("price_median_manyen")]
        records.append((body, len(group), sum(r.get("listing_count") or 0 for r in group),
                        round(st.median(medians), 1) if medians else None))
    records.sort(key=lambda r: -r[2])

    start = 4
    _header(ws, start, ["ボディタイプ", "車種数", "掲載台数", "価格中央値(万円)"])
    last = _rows(ws, start, records, {2: INT_FMT, 3: INT_FMT, 4: MANYEN_FMT})

    chart = BarChart()
    chart.type = "col"
    chart.title = "ボディタイプ別 掲載台数"
    chart.y_axis.title = "台数"
    chart.height, chart.width = 11, 24
    chart.add_data(Reference(ws, min_col=3, max_col=3, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "G4")

    price = BarChart()
    price.type = "col"
    price.title = "ボディタイプ別 価格中央値（万円）"
    price.y_axis.title = "万円"
    price.height, price.width = 11, 24
    price.add_data(Reference(ws, min_col=4, max_col=4, min_row=start, max_row=last),
                   titles_from_data=True)
    price.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(price, "G27")

    ws.column_dimensions["A"].width = 22
    for col in "BCD":
        ws.column_dimensions[col].width = 15


def model_chart(ws, compare: list[dict[str, Any]]) -> None:
    _title(ws, "掲載台数の多い車種の価格帯と値落ち",
           "価格の幅が広い車種はグレードや程度の差が大きい＝安い個体には理由がある。"
           "値落ち率が高い車種は、数年落ちを買うと得をしやすい。")

    records = [
        (f"{r.get('maker') or ''} {r.get('model_name') or ''}".strip(),
         r.get("listing_count"), r.get("price_min_manyen"),
         r.get("price_median_manyen"), r.get("price_max_manyen"),
         r.get("depreciation_pct"))
        for r in compare[:TOP_MODELS]
    ]

    start = 4
    _header(ws, start, ["車種", "掲載台数", "最安(万円)", "中央値(万円)",
                        "最高(万円)", "値落ち率(%/年)"])
    last = _rows(ws, start, records,
                 {2: INT_FMT, 3: MANYEN_FMT, 4: MANYEN_FMT, 5: MANYEN_FMT, 6: "0.0"})

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = f"掲載上位{len(records)}車種の価格帯（万円）"
    chart.x_axis.title = "万円"
    chart.height, chart.width = 18, 26
    chart.add_data(Reference(ws, min_col=3, max_col=5, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "I4")

    drop = LineChart()
    drop.title = "値落ち率（%/年）"
    drop.y_axis.title = "%"
    drop.height, drop.width = 11, 26
    drop.add_data(Reference(ws, min_col=6, max_col=6, min_row=start, max_row=last),
                  titles_from_data=True)
    drop.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(drop, "I40")

    ws.column_dimensions["A"].width = 26
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15


# -------------------------------------------------------------------- 組み立て


def _readme(ws, title: str, note: str, counts: list[tuple[str, str]],
            sheets: list[tuple[str, str]], snapshot: str | None) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 104
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    lines: list[tuple[str, str]] = [
        ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("取得時点", snapshot or "不明"),
        ("この本の位置づけ", note),
        ("", ""),
        ("■ 収録", ""),
        *counts,
        ("", ""),
        ("■ シート", ""),
        *sheets,
        ("", ""),
        ("■ 他の本", ""),
        ("souba_all.xlsx", "全2,237車種のカタログ。車種を決めていないときの索引。"),
        ("souba_minivan.xlsx", "ミニバン。深掘り20車種は口コミ・不具合・リコールまで入っている。"),
        ("souba_standard.xlsx", "乗用車（ミニバン・トラックを除く）。"),
        ("souba_classics.xlsx", "1988〜2001年式の旧車。1台ずつの在庫とヤフオク落札。"),
    ]
    for i, (key, value) in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=key)
        if key.startswith("■"):
            cell.font = TITLE_FONT
        body = ws.cell(row=i, column=2, value=value)
        body.alignment = body.alignment.copy(wrap_text=True, vertical="top")


def catalog_book(
    path: Path,
    *,
    title: str,
    note: str,
    body_types: Iterable[str] | None = None,
    makers: Iterable[str] | None = None,
) -> Path:
    """全車種クロールのデータから、比較・検討用のブックを 1 冊組む。"""
    summaries, snapshot = store.read_latest("wide_summary")
    by_year, _ = store.read_latest("wide_by_year")
    if not summaries:
        raise RuntimeError("全車種データがありません。先に `kakaku-ai wide` を実行してください。")

    if body_types is not None:
        wanted = set(body_types)
        summaries = [r for r in summaries if (r.get("body_type") or "") in wanted]
    if makers is not None:
        wanted_makers = set(makers)
        summaries = [r for r in summaries if (r.get("maker") or "") in wanted_makers]

    codes = {r["carsensor_code"] for r in summaries}
    by_year = [r for r in by_year if r["carsensor_code"] in codes]
    compare = compare_rows(summaries, by_year)

    total_listings = sum(r.get("listing_count") or 0 for r in compare)
    wb = Workbook()
    _readme(
        wb.active, title, note,
        [("車種", f"{len(compare)}車種"),
         ("掲載台数", f"{total_listings:,}台（カーセンサー）"),
         ("メーカー", f"{len({r.get('maker') for r in compare})}社"),
         ("年式別相場", f"{len(by_year)}行")],
        [("グラフ_メーカー別", "メーカーごとの車種数・掲載台数・価格中央値。"),
         ("グラフ_ボディタイプ別", "用途で絞り込むための俯瞰。"),
         ("グラフ_車種別", "掲載の多い車種の価格帯と値落ち率。"),
         ("車種比較", "1行1車種。価格帯・掲載台数・走行距離・値落ち率を横並びにしたメイン表。"
                  "メーカー・ボディタイプ・価格でフィルタして候補を絞る。"),
         ("年式別相場", "車種 × 年式のマトリクス。車種を決めたあと、どの年式が狙い目かを見る。"),
         ("年式別相場_明細", "上の元データ。掲載台数や四分位も入っている。")],
        snapshot,
    )
    wb.active.title = "README"

    maker_chart(wb.create_sheet("グラフ_メーカー別"), compare)
    body_chart(wb.create_sheet("グラフ_ボディタイプ別"), compare)
    model_chart(wb.create_sheet("グラフ_車種別"), compare)

    _write_table(wb.create_sheet("車種比較"), COMPARE_COLUMNS, compare,
                 number_formats=COMPARE_FORMATS)
    year_pivot(wb.create_sheet("年式別相場"), compare, by_year)
    _write_table(
        wb.create_sheet("年式別相場_明細"),
        [("maker", "メーカー"), ("origin", "国産/輸入"), ("body_type", "ボディタイプ"),
         ("model_name", "車種"), ("model_year", "年式"),
         ("is_open_bucket", "以前まとめ"), ("listing_count", "掲載台数"),
         ("retail_p25_manyen", "25%(万円)"), ("retail_median_manyen", "中央値(万円)"),
         ("retail_mean_manyen", "平均(万円)"), ("retail_p75_manyen", "75%(万円)"),
         ("url", "URL")],
        sorted(by_year, key=lambda r: (r.get("maker") or "", r.get("model_name") or "",
                                       r.get("model_year") or 0)),
        number_formats={"model_year": "0", "listing_count": INT_FMT,
                        "retail_p25_manyen": MANYEN_FMT, "retail_median_manyen": MANYEN_FMT,
                        "retail_mean_manyen": MANYEN_FMT, "retail_p75_manyen": MANYEN_FMT},
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("%s（%s車種 / 掲載 %s台）", path.name, len(compare), f"{total_listings:,}")
    return path
