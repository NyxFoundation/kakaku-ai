"""全スナップショットから 1 つの xlsx を組み立てる。

方針:

* 相場は **long format（1行 = スナップショット日 × 車種 × 年式）** で持つ。
  週を重ねるほど行が増えるだけで、列は増えない。
* そのうえで「推移」シートに 行 = 車種×年式 / 列 = スナップショット日 の
  ピボットを 1 枚用意しておく。折れ線を引くならここを選ぶだけで済む。
* 最新断面だけ見たい人向けに「最新」シートも別に出す。
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import aggregate, charts, running_cost, store
from .charts import _median
from .vehicles import VehicleSet, load_vehicles

log = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")

MANYEN_FMT = "#,##0.0"
INT_FMT = "#,##0"
PCT_FMT = "0.0"
MAX_COST_SERIES = 12   # 折れ線はこれ以上並べると読めない


def _write_table(
    ws: Worksheet,
    columns: Sequence[tuple[str, str]],
    rows: Iterable[dict[str, Any]],
    *,
    number_formats: dict[str, str] | None = None,
    freeze: str = "A2",
    wrap_columns: set[str] | None = None,
) -> int:
    """(キー, 見出し) の並びに従って dict の列を書き出す。"""
    number_formats = number_formats or {}
    wrap_columns = wrap_columns or set()

    for col_index, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    count = 0
    for row_index, row in enumerate(rows, start=2):
        for col_index, (key, _) in enumerate(columns, start=1):
            value = row.get(key)
            if isinstance(value, bool):
                value = "○" if value else ""
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = BORDER
            if key in number_formats:
                cell.number_format = number_formats[key]
            if key in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif row_index % 2 == 0:
                cell.fill = ALT_FILL
        count += 1

    ws.freeze_panes = freeze
    if count:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{count + 1}"

    for col_index, (key, header) in enumerate(columns, start=1):
        width = 34 if key in wrap_columns else max(10, min(24, len(header) * 2 + 4))
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[1].height = 32
    return count


# --------------------------------------------------------------------- sheets

PRICE_COLUMNS: list[tuple[str, str]] = [
    ("snapshot_date", "時点"),
    ("vehicle_name", "車種"),
    ("generation", "世代"),
    ("model_year", "年式"),
    ("auction_n", "ヤフオク\n落札件数"),
    ("auction_min_manyen", "ヤフオク\n最安(万円)"),
    ("auction_p25_manyen", "ヤフオク\n25%(万円)"),
    ("auction_median_manyen", "ヤフオク\n中央値(万円)"),
    ("auction_mean_manyen", "ヤフオク\n平均(万円)"),
    ("auction_p75_manyen", "ヤフオク\n75%(万円)"),
    ("auction_max_manyen", "ヤフオク\n最高(万円)"),
    ("auction_median_mileage_km", "落札車の\n中央走行距離(km)"),
    ("auction_unknown_repair_n", "うち修復歴\n不明の件数"),
    ("retail_n", "小売\n掲載台数"),
    ("retail_p25_manyen", "小売\n25%(万円)"),
    ("retail_median_manyen", "小売\n中央値(万円)"),
    ("retail_mean_manyen", "小売\n平均(万円)"),
    ("retail_p75_manyen", "小売\n75%(万円)"),
    ("jmty_n", "ジモティー\n掲載件数"),
    ("jmty_median_manyen", "ジモティー\n中央値(万円)"),
    ("retail_minus_auction_manyen", "小売−落札\n(万円)"),
    ("retail_premium_pct", "小売プレミアム\n(%)"),
]

PRICE_FORMATS = {
    "auction_n": INT_FMT,
    "auction_min_manyen": MANYEN_FMT,
    "auction_p25_manyen": MANYEN_FMT,
    "auction_median_manyen": MANYEN_FMT,
    "auction_mean_manyen": MANYEN_FMT,
    "auction_p75_manyen": MANYEN_FMT,
    "auction_max_manyen": MANYEN_FMT,
    "auction_median_mileage_km": INT_FMT,
    "auction_unknown_repair_n": INT_FMT,
    "retail_n": INT_FMT,
    "retail_p25_manyen": MANYEN_FMT,
    "retail_median_manyen": MANYEN_FMT,
    "retail_mean_manyen": MANYEN_FMT,
    "retail_p75_manyen": MANYEN_FMT,
    "jmty_n": INT_FMT,
    "jmty_median_manyen": MANYEN_FMT,
    "retail_minus_auction_manyen": MANYEN_FMT,
    "retail_premium_pct": PCT_FMT,
}


def _readme_sheet(
    ws: Worksheet,
    vehicles: VehicleSet,
    snapshots: list[str],
    counts: dict[str, int],
    *,
    price_date: str | None = None,
) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 110

    makers = []
    for v in vehicles:
        if v.maker not in makers:
            makers.append(v.maker)
    ws["A1"] = f"{vehicles.body_type} 中古車相場データベース（{'・'.join(makers)}）"
    ws["A1"].font = TITLE_FONT

    lines: list[tuple[str, str]] = [
        ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (
            "対象",
            f"{vehicles.body_type} {len(vehicles)}車種（{'・'.join(makers)}）/ "
            f"{vehicles.model_year_from}年式以降 / 全国",
        ),
        ("収録スナップショット", f"{len(snapshots)}件: {', '.join(snapshots)}"),
        # 深掘り20車種（週次）と全車種クロール（別サイクル）は別々の日に走るので、
        # 「最新」といっても日付が揃わない。どちらの断面かを明記しておく。
        (
            "各断面の時点",
            " / ".join(
                f"{label} {date or '未取得'}"
                for label, date in (("相場・落札:", price_date),)
            ),
        ),
        ("", ""),
        ("■ シートの読み方", ""),
        ("グラフ_価格差", "落札価格と店頭価格の差が大きい順の横棒。走行距離も併記。まずここ。"),
        ("グラフ_車種別", "車種ごとの落札相場と小売相場を並べた棒グラフ。"),
        ("グラフ_年式別", "車種 × 年式 の落札中央値マトリクス（行内で色分け）と値落ちカーブ。"),
        ("相場_最新", "直近スナップショットの 車種×年式 相場。数字で見たいときはここ。"),
        ("相場_時系列", "全スナップショットを積んだ long format。ピボットの元データ。"),
        (
            "相場_累計",
            "全スナップショットの落札を auction_id で名寄せしてから年式集計したもの。"
            "ヤフオクは 180 日ぶんしか返さないが、週を重ねるとここだけ期間が伸びて "
            "n が厚くなる。サンプルが欲しいときはこちら。",
        ),
        ("推移_落札中央値", "行=車種×年式 / 列=時点。折れ線グラフはここから引く。"),
        ("推移_小売中央値", "同上、小売（カーセンサー掲載）側。"),
        ("車種サマリ", "車種単位の価格レンジ・掲載台数・満足度など。"),
        ("口コミ_年式別", "みんカラのレビューを年式で集計したもの。"),
        ("口コミ_明細", "レビュー個票（満足点・不満点・総評）。"),
        (
            "壊れやすい点",
            "国交省に寄せられた不具合通報を 世代 × 装置 で集計。整備観点の弱点はここ。"
            "世代 =「（車種全体）」の行は車種まるごとのロールアップ。",
        ),
        (
            "不具合_明細",
            "国交省に寄せられた不具合通報 1 件ずつ。型式・初度登録年月・走行距離と"
            "症状の全文つき。「壊れやすい点」の元データ。",
        ),
        ("リコール", "国交省リコール届出。不具合装置・状況・改善措置つき。"),
        (
            "店頭_成約推定",
            "カーセンサーの在庫を個体で追い、掲載が消えたもの＝売れたとみなしたもの。"
            "店頭の成約価格はどこにも公開されていないので、これが唯一の手がかり。"
            "2 回目のクロールから貯まりはじめる。",
        ),
        (
            "落札明細",
            "ヤフオク!の落札 1 台ずつ（全スナップショットを名寄せした累計）。"
            "年式・走行距離・修復歴に加え、商品ページから取ったグレード・車検・"
            "諸費用込み総額つき。",
        ),
        (
            "年間維持費",
            "1年あたりいくらかかるか。**現金支出（税・自賠責・任意保険・車検・整備・"
            "燃料・整備予備費）はここが本体**で、車種と排気量と走行距離で決まる。\n"
            "年間走行距離は落札実績から出している（年式ごとの走行距離中央値÷車齢の"
            "中央値）。一律の仮定ではないので、ミニバンでも 1.0万〜1.5万km/年 と"
            "車種で差が出る。燃料費はここに効く。\n"
            "整備予備費は、国交省の不具合情報にある「装置ごとの発生時走行距離」を"
            "保有5年で通過するかで見積もったもの。**通報件数から発生確率は出せない**"
            "（母数＝その車種の総保有台数が分からず、遭っても通報する人はごく一部）"
            "ので、通過する装置の修理費を全部足したものに 0.5 を掛けている。"
            "この 0.5 は素の仮定。修理費そのものも相場観であって実測ではない。\n"
            "「購入時点で通過済み」の装置は、前オーナーが直していれば済んでいるが、"
            "手つかずなら遅れているだけ。**記録簿で確認すべき項目**として出している。\n"
            "リコールは入れていない。メーカー負担の無償修理なので持ち主の出費に"
            "ならない（未対策かどうかは別途「リコール」シートで確認）。\n"
            "値落ちは年式ごとの落札中央値の差から。1年式あたりの落札が3〜8件しか"
            "なく世代交代もまたぐので参考値。根拠列が「頭打ち(世代交代)」の行は、"
            "モデルチェンジの段差を値落ちとして数えないよう抑えた行。\n"
            "任意保険6万・車検基本料6万/2年・整備3万・175円/L・駐車場0円を仮定。",
        ),
        ("車種マスタ", "深掘り対象 20 車種の世代・型式の一覧。"),
        ("", ""),
        ("■ 相場の作り方", ""),
        (
            "オークション相場",
            "ヤフオク!「中古車・新車」カテゴリの終了180日間の落札から、"
            "商品ページの firstRegYear（無ければ検索結果の carSpec.modelDate）を年式として集計。"
            "メーター交換車と修復歴ありは除外し、件数は別列に残す。"
            "修復歴『わからない』は個人出品では普通の申告なので残している（列で件数が見える）。",
        ),
        (
            "小売相場",
            "カーセンサーの相場ページにある「価格帯 × 年式」度数分布から、"
            "ビン内一様分布を仮定したグループ中央値で推定。最上位ビン（420万円以上）は幅20万円と仮定。",
        ),
        (
            "小売プレミアム",
            "(小売中央値 / 落札中央値 - 1) × 100。業販と小売の価格差の目安。",
        ),
        (
            "ジモティー掲載",
            "ジモティーに出ている売り希望額の中央値。**業者と個人が混ざっている**"
            "（提携サイト＝販売店の在庫フィードが大半で、直接投稿にも業者が多い）。"
            "成約価格ではないので、落札中央値と同列には比べられない。参考値として置いてある。",
        ),
        ("", ""),
        ("■ 出典", ""),
        ("ヤフオク!", "https://auctions.yahoo.co.jp/closedsearch/closedsearch （robots.txt で Allow）"),
        ("カーセンサー", "https://www.carsensor.net/usedcar/souba/"),
        ("価格.com", "https://kakaku.com/kuruma/"),
        ("みんカラ", "https://minkara.carview.co.jp/car/toyota/"),
        ("ジモティー", "https://jmty.jp/all/car-toy/ （掲載価格。業者・個人混在）"),
        ("国土交通省", "https://renrakuda.mlit.go.jp/renrakuda/ （リコール届出情報 / 自動車不具合情報ホットライン）"),
        ("", ""),
        ("■ 注意", ""),
        (
            "サンプル数",
            "年式によっては落札件数が 1〜2 台しかない。auction_n を必ず見ること。"
            "週次で積み上がるので、時系列シート側で複数週をまとめれば精度は上がる。",
        ),
        ("価格の単位", "すべて万円（税込・諸費用別）。走行距離は km。"),
        ("更新", "毎週クロールして新しい時点の行を追記する。過去の行は書き換えない。"),
        ("リポジトリ", "https://github.com/NyxFoundation/kakaku-ai"),
        ("", ""),
        ("■ 今回のスナップショット行数", ""),
    ]
    lines += [(k, str(v)) for k, v in counts.items()]

    for i, (label, value) in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=label.startswith("■"))
        cell = ws.cell(row=i, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


COMPARE_COLUMNS: list[tuple[str, str]] = [
    ("maker", "メーカー"),
    ("vehicle_name", "車種"),
    ("production_period", "生産期間"),
    ("retail_n", "掲載台数"),
    ("shop_count", "取扱\n店舗数"),
    ("auction_n", "落札件数\n(180日)"),
    ("auction_median_manyen", "落札中央値\n(万円)"),
    ("retail_median_manyen", "店頭中央値\n(万円)"),
    ("retail_premium_pct", "店頭が高い\n(%)"),
    ("matched_years", "比較に使った\n年式数"),
    ("new_price_manyen", "新車価格\n(万円)"),
    ("depreciation_pct", "値落ち率\n(%/年)"),
    ("mileage_median_km", "掲載車の\n走行中央値(km)"),
    ("review_score", "口コミ\n総合"),
    ("review_count", "口コミ\n件数"),
    ("review_loading", "積載性"),
    ("review_running_cost", "燃費"),
    ("review_comfort", "快適性"),
    ("defect_n", "不具合\n通報数"),
    ("defect_top", "不具合の\n最多装置"),
    ("recall_n", "リコール\n件数"),
]

COMPARE_FORMATS = {
    "retail_n": INT_FMT, "shop_count": INT_FMT, "auction_n": INT_FMT,
    "auction_median_manyen": MANYEN_FMT, "retail_median_manyen": MANYEN_FMT,
    "retail_premium_pct": PCT_FMT, "matched_years": INT_FMT,
    "new_price_manyen": MANYEN_FMT,
    "depreciation_pct": PCT_FMT, "mileage_median_km": INT_FMT,
    "review_score": "0.0", "review_count": INT_FMT,
    "review_loading": "0.0", "review_running_cost": "0.0", "review_comfort": "0.0",
    "defect_n": INT_FMT, "recall_n": INT_FMT,
}


def _cost_chart_sheet(ws: Worksheet, sim_rows: list[dict[str, Any]]) -> None:
    """横軸=何年目 / 縦軸=維持費の累計 の折れ線。系列は 車種 × 年式。

    段差が意味を持つグラフなので、均さずに年ごとの実額を積んでいる。
    2 年に 1 回の車検、13年目・18年目の増税、修理が来た年がそのまま
    折れ線の傾きの変化として出る。

    系列が多すぎると読めないので、10 年目の累計が安い順に絞る。
    """
    series: dict[str, dict[int, float]] = {}
    meta: dict[str, dict[str, Any]] = {}
    for row in sim_rows:
        key = f"{row['vehicle_name']} {row['model_year']}年式"
        series.setdefault(key, {})[row["year"]] = row["cumulative_manyen"]
        meta[key] = row
    if not series:
        return

    years = sorted({r["year"] for r in sim_rows})
    ordered = sorted(series, key=lambda k: series[k].get(years[-1], 0))[:MAX_COST_SERIES]

    charts._title(
        ws,
        f"買ってから何年目までに、累計いくらかかるか（万円）",
        "車両価格・税・自賠責・任意保険・車検・整備・燃料・修理・値落ちの累計。"
        "2年に1回の車検、13年目と18年目の増税、修理が来た年で傾きが変わる。"
        "安い順に上位%s系列。値落ちと修理は仮定を含むので、順位の目安として読むこと。"
        % len(ordered),
    )

    start = 4
    charts._header(ws, start, ["経過年", *ordered])
    for offset, year in enumerate(years, start=1):
        ws.cell(row=start + offset, column=1, value=year).border = BORDER
        for col, key in enumerate(ordered, start=2):
            cell = ws.cell(row=start + offset, column=col, value=series[key].get(year))
            cell.number_format = MANYEN_FMT
            cell.border = BORDER
    last = start + len(years)

    chart = LineChart()
    chart.title = "維持費の累計（万円）"
    chart.y_axis.title = "累計（万円）"
    chart.x_axis.title = "経過年"
    chart.height, chart.width = 16, 30
    chart.add_data(Reference(ws, min_col=2, max_col=1 + len(ordered),
                             min_row=start, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, f"A{last + 3}")

    ws.column_dimensions["A"].width = 10
    for col in range(2, len(ordered) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.row_dimensions[start].height = 40


def _compare_rows(
    vehicles: VehicleSet,
    price_latest: list[dict[str, Any]],
    summary: list[dict[str, Any]],
    defects: list[dict[str, Any]],
    recalls: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """1 行 1 車種の比較表。**この 1 枚で候補を絞れる**ことを狙う。

    相場・口コミ・整備上の弱点をそれぞれ別のシートに散らすと、20 車種を
    突き合わせるのに何度も行き来することになる。決めるのに要る数字だけを
    ここに集めて、残りは詳細シートに置く。
    """
    by_key = {r["vehicle_key"]: r for r in summary}
    years: dict[str, list[dict[str, Any]]] = {}
    for row in price_latest:
        years.setdefault(row["vehicle_key"], []).append(row)

    defect_n: dict[str, int] = {}
    defect_top: dict[str, str] = {}
    for row in defects:
        key = row["vehicle_key"]
        defect_n[key] = defect_n.get(key, 0) + (row.get("report_count") or 0)
    for key in defect_n:
        top = max((r for r in defects if r["vehicle_key"] == key),
                  key=lambda r: r.get("report_count") or 0, default=None)
        if top:
            defect_top[key] = f"{top['defective_device']}（{top['report_count']}件）"

    recall_n: dict[str, int] = {}
    for row in recalls:
        recall_n[row["vehicle_key"]] = recall_n.get(row["vehicle_key"], 0) + 1

    out: list[dict[str, Any]] = []
    for vehicle in vehicles:
        rows = years.get(vehicle.key, [])
        info = by_key.get(vehicle.key, {})

        auctions = [(r["model_year"], r["auction_median_manyen"], r.get("auction_n") or 0)
                    for r in rows if r.get("auction_median_manyen")]

        # 落札と店頭を比べるときは**両方そろっている年式だけ**で中央値を取る。
        # 片方にしか無い年式を混ぜると、落札は古い年式に、店頭は新しい年式に
        # 寄っている車種で差が実態の倍以上に出る（フリードで +186% になっていた）
        both = [(r["model_year"], r["auction_median_manyen"], r["retail_median_manyen"])
                for r in rows
                if r.get("auction_median_manyen") and r.get("retail_median_manyen")]
        auction_median = _median([b[1] for b in both])
        retail_median = _median([b[2] for b in both])
        matched_years = len(both)
        # 差は「年式ごとの比」の中央値で取る。中央値どうしを割ると、2 つの列を
        # それぞれ独立に並べ替えたことになって別々の年式が突き合わされる
        # （フリードで 2018年の落札 と 2019年の店頭 を比べていた）
        premium = _median([b[2] / b[1] for b in both if b[1]])

        # 値落ちは 落札中央値 の最古年式と最新年式から。小売は掲載側の
        # 価格付けが入るので、実際に売れた値である落札のほうを使う
        drop = None
        if len(auctions) >= 2:
            auctions.sort()
            (old_y, old_p, _), (new_y, new_p, _) = auctions[0], auctions[-1]
            if new_y > old_y and new_p:
                drop = round((1 - old_p / new_p) / (new_y - old_y) * 100, 1)

        new_price = info.get("kakaku_new_price_min_manyen")
        out.append({
            "maker": vehicle.maker,
            "vehicle_name": vehicle.name,
            "production_period": info.get("carsensor_production_period"),
            "retail_n": info.get("carsensor_listing_count"),
            "shop_count": info.get("carsensor_shop_count"),
            "auction_n": sum(a[2] for a in auctions),
            "auction_median_manyen": auction_median,
            "retail_median_manyen": retail_median,
            "matched_years": matched_years,
            "retail_premium_pct": round((premium - 1) * 100, 1) if premium else None,
            "new_price_manyen": new_price,
            "depreciation_pct": drop,
            "mileage_median_km": info.get("carsensor_retail_median_mileage_km"),
            "review_score": info.get("carsensor_review_score_overall"),
            "review_count": info.get("carsensor_review_count"),
            "review_loading": info.get("carsensor_review_loading"),
            "review_running_cost": info.get("carsensor_review_running_cost"),
            "review_comfort": info.get("carsensor_review_comfort"),
            "defect_n": defect_n.get(vehicle.key),
            "defect_top": defect_top.get(vehicle.key),
            "recall_n": recall_n.get(vehicle.key),
        })
    out.sort(key=lambda r: -(r.get("retail_n") or 0))
    return out


COST_COLUMNS: list[tuple[str, str]] = [
    ("vehicle_name", "車種"),
    ("model_year", "年式"),
    ("age", "車齢"),
    ("displacement_l", "排気量\n(L)"),
    ("price_manyen", "車両価格\n(万円)"),
    ("vehicle_tax_yen", "自動車税"),
    ("weight_tax_yen", "重量税\n(年割)"),
    ("compulsory_insurance_yen", "自賠責\n(年割)"),
    ("voluntary_insurance_yen", "任意保険"),
    ("inspection_yen", "車検\n(年割)"),
    ("maintenance_yen", "整備・消耗品"),
    ("parking_yen", "駐車場"),
    ("annual_km", "年間走行\n(km)"),
    ("annual_km_source", "走行距離の\n出どころ"),
    ("fuel_yen", "燃料"),
    ("fuel_economy_kml", "実燃費\n(km/L)"),
    ("odometer_km", "購入時\n走行(km)"),
    ("odometer_end_km", "5年後\n走行(km)"),
    ("repair_devices", "保有中に来る\n故障（距離帯）"),
    ("repair_passed", "購入時点で\n通過済み"),
    ("repair_worst_yen", "修理費\n全部起きたら"),
    ("repair_reserve_yen", "整備予備費\n/年"),
    ("cash_out_yen", "現金支出\n合計/年"),
    ("depreciation_yen", "値落ち\n(参考)"),
    ("depreciation_basis", "値落ちの\n根拠"),
    ("total_yen", "総額/年"),
    ("monthly_yen", "月あたり"),
    ("is_old_car", "13年超\n重課"),
]

COST_FORMATS = {
    "model_year": "0", "age": "0", "displacement_l": "0.0",
    "price_manyen": MANYEN_FMT, "fuel_economy_kml": "0.0",
    **{k: INT_FMT for k in (
        "vehicle_tax_yen", "weight_tax_yen", "compulsory_insurance_yen",
        "voluntary_insurance_yen", "inspection_yen", "maintenance_yen",
        "parking_yen", "fuel_yen", "cash_out_yen", "depreciation_yen",
        "total_yen", "monthly_yen", "annual_km", "odometer_km",
        "odometer_end_km", "repair_worst_yen", "repair_reserve_yen")},
}


def _pivot(
    ws: Worksheet,
    rows: list[dict[str, Any]],
    snapshots: list[str],
    value_key: str,
    value_label: str,
) -> None:
    """行 = 車種×年式 / 列 = スナップショット日 のピボット。"""
    index: dict[tuple[str, Any, str], dict[str, Any]] = {}
    for row in rows:
        key = (row.get("vehicle_name", ""), row.get("model_year"), row.get("generation", ""))
        index.setdefault(key, {})[row["snapshot_date"]] = row.get(value_key)

    headers = ["車種", "世代", "年式"] + snapshots + ["初回比(%)"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER

    row_index = 2
    for (name, year, generation) in sorted(
        index, key=lambda k: (k[0], -(k[1] or 0))
    ):
        series = index[(name, year, generation)]
        ws.cell(row=row_index, column=1, value=name).border = BORDER
        ws.cell(row=row_index, column=2, value=generation).border = BORDER
        ws.cell(row=row_index, column=3, value=year).border = BORDER
        values: list[float] = []
        for offset, snap in enumerate(snapshots):
            value = series.get(snap)
            cell = ws.cell(row=row_index, column=4 + offset, value=value)
            cell.number_format = MANYEN_FMT
            cell.border = BORDER
            if isinstance(value, (int, float)):
                values.append(float(value))
        if len(values) >= 2 and values[0]:
            change = ws.cell(
                row=row_index,
                column=4 + len(snapshots),
                value=round((values[-1] / values[0] - 1) * 100, 1),
            )
            change.number_format = PCT_FMT
            change.border = BORDER
        row_index += 1

    ws.freeze_panes = "D2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    for offset in range(len(snapshots) + 1):
        ws.column_dimensions[get_column_letter(4 + offset)].width = 13

    # スナップショットが 2 つ以上たまったら折れ線を置く。
    # 1 つしかないうちは点しか描けないので出さない。
    last_row = row_index - 1
    if len(snapshots) >= 2 and last_row >= 2:
        rows_to_plot = min(last_row - 1, 20)
        chart = LineChart()
        # 取得できなかった週は空セルになる。そこで線を切らずにまたいで結ぶ。
        chart.display_blanks = "span"
        chart.title = f"{value_label}の推移（上位{rows_to_plot}行）"
        chart.y_axis.title = "万円"
        chart.x_axis.title = "時点"
        chart.height = 11
        chart.width = 24
        data = Reference(
            ws,
            min_col=4,
            max_col=3 + len(snapshots),
            min_row=1,
            max_row=1 + rows_to_plot,
        )
        chart.add_data(data, titles_from_data=False, from_rows=True)
        for series, row_offset in zip(chart.series, range(2, 2 + rows_to_plot)):
            name = f"{ws.cell(row=row_offset, column=1).value} {ws.cell(row=row_offset, column=3).value}"
            series.tx = SeriesLabel(v=name)
        ws.add_chart(chart, f"{get_column_letter(6 + len(snapshots))}2")

    log.info("  pivot %s: %s行", value_label, row_index - 2)


def build(output: Path, *, vehicles: VehicleSet | None = None) -> Path:
    vehicles = vehicles or load_vehicles()
    snapshots = store.list_snapshots()
    if not snapshots:
        raise RuntimeError("スナップショットが 1 つもありません。先に crawl を実行してください。")
    # 「最新のスナップショット日」ではなく**データセットごと**に最新を取る。
    # 全車種クロールや `--sources` を絞った実行はその日の一部しか書かないので、
    # 日付で揃えると撮っていないシートが軒並み空になる。
    price_all = store.read_all("price_by_year")
    price_latest, price_date = store.read_latest("price_by_year")
    summary_latest, _ = store.read_latest("vehicle_summary")
    reviews_latest, _ = store.read_latest("reviews")
    review_summary_latest, _ = store.read_latest("review_summary")
    defects_latest, _ = store.read_latest("defect_summary")
    defect_details, _ = store.read_latest("defects")
    recalls_latest, _ = store.read_latest("recalls")
    listings_latest, _ = store.read_latest("auction_listings")
    # 落札は「終了180日間」しか取れないので、全スナップショットを auction_id で
    # 名寄せして 1 本にする。週を重ねるほど実効期間が伸び、年式ごとの n が増える。
    listings_pool = store.pooled_auction_listings()
    jmty_latest, _ = store.read_latest("jmty_listings")
    delisted, _ = store.read_latest("carsensor_delisted")

    wb = Workbook()
    counts = {
        "相場_時系列": len(price_all),
        "相場_最新": len(price_latest),
        "口コミ_明細": len(reviews_latest),
        "壊れやすい点": len(defects_latest),
        "不具合_明細": len(defect_details),
        "リコール": len(recalls_latest),
        "落札明細_累計": len(listings_pool),
        "落札明細（最新断面のみ）": len(listings_latest),
        "参考_ジモティー掲載": len(jmty_latest),
        "店頭_成約推定": len(delisted),
    }

    _readme_sheet(wb.active, vehicles, snapshots, counts,
                  price_date=price_date)
    wb.active.title = "README"

    # --- グラフ（断面）。表より先に置いて、開いてすぐ絵が見えるようにする ---
    charts.build(wb, price_latest, listings_pool, summary_latest, vehicles.model_year_from)

    # --- 車種比較: この 1 枚で候補を絞れるようにする ---
    _write_table(
        wb.create_sheet("車種比較"),
        COMPARE_COLUMNS,
        _compare_rows(vehicles, price_latest, summary_latest, defects_latest, recalls_latest),
        number_formats=COMPARE_FORMATS,
        wrap_columns={"defect_top"},
    )

    # --- グラフ_維持費: 何年目までにいくらか ---
    sim_rows = running_cost.simulate_table(listings_pool, price_latest,
                                           defects=defects_latest)
    if sim_rows:
        _cost_chart_sheet(wb.create_sheet("グラフ_維持費"), sim_rows)

    # --- 年間維持費 ---
    cost_rows = running_cost.build_table(listings_pool, price_latest,
                                         defects=defects_latest)
    if cost_rows:
        ws = wb.create_sheet("年間維持費")
        _write_table(ws, COST_COLUMNS, cost_rows, number_formats=COST_FORMATS,
                     wrap_columns={"depreciation_basis", "repair_devices",
                                   "repair_passed"})

    # --- 車種マスタ ---
    ws = wb.create_sheet("車種マスタ")
    master_rows = []
    for v in vehicles:
        for gen in v.generations:
            master_rows.append(
                {
                    "maker": v.maker,
                    "vehicle_name": v.name,
                    "vehicle_key": v.key,
                    "generation": gen.code,
                    "period": gen.label.split(" ", 1)[-1].strip("()"),
                    "models": ", ".join(gen.models),
                    "carsensor": ", ".join(v.carsensor_codes),
                    "kakaku": v.kakaku_item_id or "",
                    "minkara": v.minkara_slug or "",
                    "yahoo": ", ".join(str(c) for c in v.yahoo_categories),
                }
            )
    _write_table(
        ws,
        [
            ("maker", "メーカー"),
            ("vehicle_name", "車種"),
            ("vehicle_key", "キー"),
            ("generation", "世代"),
            ("period", "販売期間"),
            ("models", "型式"),
            ("carsensor", "カーセンサー\nコード"),
            ("kakaku", "価格.com\nID"),
            ("minkara", "みんカラ\nslug"),
            ("yahoo", "ヤフオク\nカテゴリ"),
        ],
        master_rows,
        wrap_columns={"models"},
    )

    # --- 相場 ---
    _write_table(
        wb.create_sheet("相場_最新"),
        [c for c in PRICE_COLUMNS if c[0] != "snapshot_date"],
        price_latest,
        number_formats=PRICE_FORMATS,
    )
    _write_table(
        wb.create_sheet("相場_時系列"),
        PRICE_COLUMNS,
        price_all,
        number_formats=PRICE_FORMATS,
    )

    # --- 相場_累計: 全スナップショットの落札を名寄せしてから年式集計 ---
    # 断面ごとの n が少ない年式でも、週を重ねればここは厚くなっていく。
    cumulative: list[dict[str, Any]] = []
    retail_latest = {
        (r["vehicle_key"], r["model_year"]): r for r in price_latest
    }
    by_key: dict[str, list[dict[str, Any]]] = {}
    for row in listings_pool:
        by_key.setdefault(row.get("vehicle_key", ""), []).append(row)
    for vehicle in vehicles:
        rows = by_key.get(vehicle.key)
        if not rows:
            continue
        ends = sorted(r["end_time"] for r in rows if r.get("end_time"))
        window = f"{ends[0][:10]}〜{ends[-1][:10]}" if ends else ""
        for stat in aggregate.yahoo_by_year(rows, vehicle, "累計"):
            if stat["model_year"] < vehicles.model_year_from:
                continue
            retail = retail_latest.get((vehicle.key, stat["model_year"]), {})
            retail_median = retail.get("retail_median_manyen")
            auction_median = stat["auction_median_manyen"]
            stat["retail_n"] = retail.get("retail_n")
            stat["retail_median_manyen"] = retail_median
            stat["retail_premium_pct"] = (
                round((retail_median / auction_median - 1) * 100, 1)
                if retail_median and auction_median
                else None
            )
            stat["window"] = window
            cumulative.append(stat)

    _write_table(
        wb.create_sheet("相場_累計"),
        [
            ("vehicle_name", "車種"),
            ("generation", "世代"),
            ("model_year", "年式"),
            ("auction_n", "落札件数\n(累計)"),
            ("auction_min_manyen", "最安(万円)"),
            ("auction_p25_manyen", "25%(万円)"),
            ("auction_median_manyen", "中央値(万円)"),
            ("auction_mean_manyen", "平均(万円)"),
            ("auction_p75_manyen", "75%(万円)"),
            ("auction_max_manyen", "最高(万円)"),
            ("auction_median_mileage_km", "中央走行距離(km)"),
            ("unknown_repair_n", "うち修復歴\n不明"),
            ("excluded_n", "除外件数\n(メーター交換/修復歴あり)"),
            ("retail_n", "小売\n掲載台数"),
            ("retail_median_manyen", "小売中央値\n(万円・最新)"),
            ("retail_premium_pct", "小売プレミアム\n(%)"),
            ("window", "落札の対象期間"),
        ],
        cumulative,
        number_formats={
            k: (INT_FMT if k in ("auction_n", "auction_median_mileage_km", "excluded_n", "retail_n")
                else MANYEN_FMT)
            for k in PRICE_FORMATS
        } | {"retail_premium_pct": PCT_FMT, "auction_n": INT_FMT,
             "excluded_n": INT_FMT, "retail_n": INT_FMT,
             "auction_median_mileage_km": INT_FMT},
    )
    _pivot(wb.create_sheet("推移_落札中央値"), price_all, snapshots, "auction_median_manyen", "落札中央値")
    _pivot(wb.create_sheet("推移_小売中央値"), price_all, snapshots, "retail_median_manyen", "小売中央値")

    # --- 車種サマリ ---
    _write_table(
        wb.create_sheet("車種サマリ"),
        [
            ("vehicle_name", "車種"),
            ("generations", "世代"),
            ("kakaku_new_price_min_manyen", "新車価格\n下限(万円)"),
            ("kakaku_new_price_max_manyen", "新車価格\n上限(万円)"),
            ("carsensor_retail_price_min_manyen", "中古下限\n(万円)"),
            ("carsensor_retail_price_max_manyen", "中古上限\n(万円)"),
            ("carsensor_listing_count", "掲載台数\n(カーセンサー)"),
            ("carsensor_shop_count", "取扱店舗数"),
            ("kakaku_used_listing_count", "掲載台数\n(価格.com)"),
            ("carsensor_review_score_overall", "総合評価\n(カーセンサー)"),
            ("carsensor_review_count", "口コミ件数\n(カーセンサー)"),
            ("kakaku_satisfaction_score", "満足度\n(価格.com)"),
            ("kakaku_review_count", "レビュー数\n(価格.com)"),
            ("kakaku_bbs_post_count", "クチコミ数\n(価格.com)"),
            ("carsensor_review_design", "デザイン"),
            ("carsensor_review_driving", "走行性"),
            ("carsensor_review_comfort", "居住性"),
            ("carsensor_review_handling", "運転しやすさ"),
            ("carsensor_review_loading", "積載性"),
            ("carsensor_review_running_cost", "維持費"),
            ("carsensor_ranking_position", "ミニバン\nランキング"),
        ],
        summary_latest,
        number_formats={
            "carsensor_listing_count": INT_FMT,
            "kakaku_used_listing_count": INT_FMT,
            "carsensor_shop_count": INT_FMT,
            "carsensor_review_count": INT_FMT,
            "kakaku_review_count": INT_FMT,
            "kakaku_bbs_post_count": INT_FMT,
        },
        wrap_columns={"generations"},
    )

    # --- 口コミ ---
    _write_table(
        wb.create_sheet("口コミ_年式別"),
        [
            ("vehicle_name", "車種"),
            ("generation", "世代"),
            ("model_year", "年式"),
            ("review_count", "件数"),
            ("score_overall", "おすすめ度"),
            ("score_design", "デザイン"),
            ("score_driving", "走行性能"),
            ("score_ride", "乗り心地"),
            ("score_loading", "積載性"),
            ("score_fuel_economy", "燃費"),
            ("score_price", "価格"),
            ("good_points", "満足している点"),
            ("bad_points", "不満な点"),
        ],
        review_summary_latest,
        wrap_columns={"good_points", "bad_points"},
    )
    _write_table(
        wb.create_sheet("口コミ_明細"),
        [
            ("vehicle_name", "車種"),
            ("model_year", "年式"),
            ("grade", "グレード"),
            ("review_date", "レビュー日"),
            ("usage", "使用目的"),
            ("score_overall", "おすすめ度"),
            ("score_design", "デザイン"),
            ("score_driving", "走行性能"),
            ("score_ride", "乗り心地"),
            ("score_loading", "積載性"),
            ("score_fuel_economy", "燃費"),
            ("review_title", "タイトル"),
            ("good_points", "満足している点"),
            ("bad_points", "不満な点"),
            ("overall_comment", "総評"),
            ("review_url", "URL"),
        ],
        reviews_latest,
        wrap_columns={"good_points", "bad_points", "overall_comment"},
    )

    # --- 故障・リコール ---
    _write_table(
        wb.create_sheet("壊れやすい点"),
        [
            ("vehicle_name", "車種"),
            ("generation", "世代"),
            ("defective_device", "不具合装置"),
            ("report_count", "通報件数"),
            ("share_pct", "構成比(%)"),
            ("median_mileage_km", "発生時の\n中央走行距離(km)"),
            ("model_year_min", "対象年式\n最古"),
            ("model_year_max", "対象年式\n最新"),
            ("model_codes", "型式"),
            ("examples", "代表事例（直近3件）"),
        ],
        defects_latest,
        number_formats={"report_count": INT_FMT, "median_mileage_km": INT_FMT, "share_pct": PCT_FMT},
        wrap_columns={"examples", "model_codes"},
    )
    _write_table(
        wb.create_sheet("不具合_明細"),
        [
            ("vehicle_name", "車種"),
            ("generation", "世代"),
            ("model_year", "年式"),
            ("model_code", "型式"),
            ("defective_device", "不具合装置"),
            ("reception_date", "受付日"),
            ("first_registration", "初度登録"),
            ("mileage_km", "走行距離(km)"),
            ("emergence_time", "発生時期"),
            ("engine_model", "原動機型式"),
            ("prefecture", "都道府県"),
            ("summary", "症状"),
            ("control_no", "管理番号"),
        ],
        sorted(
            defect_details,
            key=lambda r: (r.get("vehicle_name", ""), r.get("reception_date") or ""),
            reverse=False,
        ),
        number_formats={"mileage_km": INT_FMT},
        wrap_columns={"summary"},
    )

    _write_table(
        wb.create_sheet("リコール"),
        [
            ("vehicle_name", "車種"),
            ("notification_date", "届出日"),
            ("notification_no", "届出番号"),
            ("defective_device", "不具合装置"),
            ("target_units", "対象台数"),
            ("models", "型式"),
            ("production_from", "製作期間\n開始"),
            ("production_to", "製作期間\n終了"),
            ("situation", "不具合の状況"),
            ("measures", "改善措置"),
        ],
        sorted(recalls_latest, key=lambda r: (r.get("vehicle_name", ""), r.get("notification_date") or ""), reverse=False),
        number_formats={"target_units": INT_FMT},
        wrap_columns={"situation", "measures", "models"},
    )

    # --- 落札明細（累計） ---
    _write_table(
        wb.create_sheet("落札明細"),
        [
            ("vehicle_name", "車種"),
            ("model_year", "年式"),
            ("generation", "世代"),
            ("grade", "グレード"),
            ("price", "落札価格(円)"),
            ("total_price", "諸費用込み\n総額(円)"),
            ("bid_count", "入札数"),
            ("mileage_km", "走行距離(km)"),
            ("mileage_type", "距離区分"),
            ("repair_type", "修復歴"),
            ("inspection_until", "車検"),
            ("transmission", "ミッション"),
            ("fuel", "燃料"),
            ("color", "色"),
            ("end_time", "終了日時"),
            ("first_seen_snapshot", "初出\nスナップショット"),
            ("title", "商品名"),
            ("url", "URL"),
        ],
        sorted(
            listings_pool,
            key=lambda r: (r.get("vehicle_name", ""), -(r.get("model_year") or 0),
                           r.get("end_time") or ""),
        ),
        number_formats={
            "price": INT_FMT, "total_price": INT_FMT, "mileage_km": INT_FMT,
        },
        wrap_columns={"title"},
    )

    # --- 店頭_成約推定（掲載が消えた個体） ---
    ws = wb.create_sheet("店頭_成約推定")
    if delisted:
        _write_table(
            ws,
            [
                ("vehicle_name", "車種"),
                ("model_year", "年式"),
                ("generation", "世代"),
                ("delisted_on", "掲載が\n消えた日"),
                ("first_seen", "初めて\n見た日"),
                ("first_price_manyen", "初回の\n総額(万円)"),
                ("last_price_manyen", "最後の\n総額(万円)"),
                ("price_cut_manyen", "値下げ幅\n(万円)"),
                ("mileage_km", "走行距離(km)"),
                ("repair_history", "修復歴"),
                ("inspection", "車検"),
                ("listing_id", "掲載ID"),
                ("url", "URL"),
            ],
            delisted,
            number_formats={
                "first_price_manyen": MANYEN_FMT,
                "last_price_manyen": MANYEN_FMT,
                "price_cut_manyen": MANYEN_FMT,
                "mileage_km": INT_FMT,
            },
        )
    else:
        ws.column_dimensions["A"].width = 100
        ws["A1"] = "店頭の成約推定（掲載が消えた個体）"
        ws["A1"].font = TITLE_FONT
        for i, line in enumerate(
            [
                "店頭でいくらで売れたかは、どのサイトも公開していない。",
                "そこでカーセンサーの在庫を個体 ID で毎週追いかけ、"
                "先週まであった掲載が消えたら「売れた」とみなして、そのときの掲載価格を記録する。",
                "",
                "いまは 1 回目の観測を取ったところなので、まだ空。次回のクロールから貯まりはじめる。",
                "",
                "読むときの注意: 掲載終了は取り下げや掲載期限切れでも起きるし、"
                "実際の成約額は値引きぶん掲載価格より下がるのが普通。あくまで上限の目安。",
            ],
            start=3,
        ):
            cell = ws.cell(row=i, column=1, value=line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- 参考: ジモティー掲載明細 ---
    _write_table(
        wb.create_sheet("参考_ジモティー掲載"),
        [
            ("vehicle_name", "車種"),
            ("model_year", "年式"),
            ("generation", "世代"),
            ("asking_price", "掲載価格(円)"),
            ("mileage_km", "走行距離(km)"),
            ("region", "地域"),
            ("is_alliance", "提携サイト\n(販売店フィード)"),
            ("looks_like_dealer", "業者ワード\nあり"),
            ("title", "タイトル"),
            ("url", "URL"),
        ],
        sorted(
            jmty_latest,
            key=lambda r: (r.get("vehicle_name", ""), -(r.get("model_year") or 0)),
        ),
        number_formats={"asking_price": INT_FMT, "mileage_km": INT_FMT},
        wrap_columns={"title"},
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    log.info("xlsx を書き出しました: %s (%s シート)", output, len(wb.sheetnames))
    return output
