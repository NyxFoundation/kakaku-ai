"""xlsx に「一目で分かる」グラフシートを足す。

スナップショットが 1 つしかなくても意味を持つ、断面のグラフだけをここに置く。
（時系列の折れ線は `推移_*` シート側で、2 断面目から自動で入る）

3 枚:

* `グラフ_車種別`   … 車種ごとの落札相場と小売相場を並べた棒グラフ
* `グラフ_年式別`   … 車種 × 年式 の落札中央値マトリクス（行内で色分け）+ 値落ちカーブ
* `グラフ_価格差`   … 落札価格と店頭価格の差が大きい順の横棒。走行距離も併記
"""

from __future__ import annotations

import logging
from typing import Any, Iterable

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
NOTE_FONT = Font(size=9, color="595959")
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MANYEN_FMT = "#,##0.0"
INT_FMT = "#,##0"
PCT_FMT = "0.0"

MANYEN = 10_000
MIN_SAMPLES = 3  # これ未満の年式は「狙い目」ランキングに載せない

# 行内での相対的な安さを色で見せる（薄い＝その車種のなかでは安い年式）
ROW_SCALE = ColorScaleRule(
    start_type="min",
    start_color="FFFFFF",
    mid_type="percentile",
    mid_value=50,
    mid_color="BDD7EE",
    end_type="max",
    end_color="2E75B6",
)


def _median(values: list[float]) -> float | None:
    if not values:
        return None
    values = sorted(values)
    mid = len(values) // 2
    if len(values) % 2:
        return values[mid]
    return (values[mid - 1] + values[mid]) / 2


def _weighted_median(pairs: list[tuple[float, int]]) -> float | None:
    """(値, 重み) の列から重み付き中央値。

    小売相場は年式ごとの中央値と掲載台数しか持っていないので、
    車種まるごとの中央値はこれで近似する。
    """
    pairs = [(v, w) for v, w in pairs if v is not None and w]
    if not pairs:
        return None
    pairs.sort(key=lambda x: x[0])
    total = sum(w for _, w in pairs)
    cumulative = 0
    for value, weight in pairs:
        cumulative += weight
        if cumulative >= total / 2:
            return value
    return pairs[-1][0]


def _header(ws: Worksheet, row: int, labels: Iterable[str]) -> None:
    for col_index, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col_index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _title(ws: Worksheet, text: str, note: str) -> None:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    ws["A2"] = note
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = Alignment(vertical="top")


def _clean_listings(listings: list[dict[str, Any]], model_year_from: int) -> list[dict[str, Any]]:
    """実走行・修復歴なし・対象年式以降の落札だけを残す。"""
    return [
        r
        for r in listings
        if (r.get("model_year") or 0) >= model_year_from
        and r.get("mileage_type") in (None, "REAL_MILEAGE")
        and r.get("repair_type") in (None, "NONE")
        and r.get("price")
    ]


# --------------------------------------------------------------- 1. 車種別


def vehicle_sheet(
    ws: Worksheet,
    price_rows: list[dict[str, Any]],
    listings: list[dict[str, Any]],
    retail_mileage: dict[str, Any],
    model_year_from: int,
) -> None:
    _title(
        ws,
        f"車種別の相場（{model_year_from}年式以降）",
        "落札中央値はヤフオク!の落札明細から実測。小売中央値は、**落札と同じ年式構成に揃えて** "
        "カーセンサーの年式別中央値を重み付けしたもの。"
        "単純に掲載台数で重み付けすると、落札は古い年式・掲載は新しい年式に偏っているせいで"
        "差額が実態より大きく出てしまうため、こう揃えている。"
        "それでも落札車は走行距離が長く保証も付かないので、差額がそのまま利ざやではない。",
    )

    by_vehicle: dict[str, dict[str, Any]] = {}
    for row in _clean_listings(listings, model_year_from):
        slot = by_vehicle.setdefault(
            row["vehicle_name"], {"prices": [], "mileage": [], "years": {}}
        )
        slot["prices"].append(row["price"] / MANYEN)
        slot["years"][row["model_year"]] = slot["years"].get(row["model_year"], 0) + 1
        if row.get("mileage_km"):
            slot["mileage"].append(row["mileage_km"])

    # 車種 → {年式: 小売中央値} と 掲載台数合計
    retail_by_year: dict[str, dict[int, float]] = {}
    listing_total: dict[str, int] = {}
    for row in price_rows:
        name = row["vehicle_name"]
        if row.get("retail_n"):
            listing_total[name] = listing_total.get(name, 0) + row["retail_n"]
        if row.get("retail_median_manyen"):
            retail_by_year.setdefault(name, {})[row["model_year"]] = row["retail_median_manyen"]

    names = sorted(
        set(by_vehicle) | set(retail_by_year),
        key=lambda n: -(_median(by_vehicle.get(n, {}).get("prices", [])) or 0),
    )

    start = 4
    _header(
        ws,
        start,
        ["車種", "落札中央値\n(万円)", "小売中央値\n(万円)\n※落札と同じ年式構成",
         "小売−落札\n(万円)", "落札件数", "掲載台数", "落札車の\n中央走行距離(km)",
         "掲載車の\n中央走行距離(km)", "対象年式"],
    )

    for offset, name in enumerate(names, start=1):
        slot = by_vehicle.get(name, {})
        auction = _median(slot.get("prices", []))
        mileage = _median([float(m) for m in slot.get("mileage", [])])

        # 落札の年式ヒストグラムを重みにして、小売側を同じ構成に揃える
        year_counts: dict[int, int] = slot.get("years", {})
        per_year = retail_by_year.get(name, {})
        matched = [(per_year[y], c) for y, c in year_counts.items() if y in per_year]
        retail_median = _weighted_median(matched)
        covered = sorted(y for y, _ in ((y, c) for y, c in year_counts.items() if y in per_year))

        spread = round(retail_median - auction, 1) if auction and retail_median else None
        values = [
            name,
            round(auction, 1) if auction else None,
            round(retail_median, 1) if retail_median else None,
            spread,
            len(slot.get("prices", [])) or None,
            listing_total.get(name),
            int(mileage) if mileage else None,
            retail_mileage.get(name),
            f"{covered[0]}〜{covered[-1]}年" if covered else "",
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=start + offset, column=col_index, value=value)
            cell.border = BORDER
            if col_index in (2, 3, 4):
                cell.number_format = MANYEN_FMT
            elif col_index in (5, 6, 7, 8):
                cell.number_format = INT_FMT

    last = start + len(names)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = f"車種別 落札相場 vs 小売相場（{model_year_from}年式以降・年式構成を揃えた中央値）"
    chart.y_axis.title = "万円"
    chart.height = 12
    chart.width = 28
    chart.add_data(
        Reference(ws, min_col=2, max_col=3, min_row=start, max_row=last),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "K4")

    ws.column_dimensions["A"].width = 18
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 15
    ws.row_dimensions[start].height = 44


# --------------------------------------------------------------- 2. 年式別


def year_sheet(
    ws: Worksheet, listings: list[dict[str, Any]], model_year_from: int
) -> None:
    _title(
        ws,
        f"車種 × 年式 の落札相場（{model_year_from}年式以降）",
        "セルの色は「その車種のなかでの相対的な高さ」。薄いほどその車種では安い年式。"
        "下段は落札件数で、件数が 1〜2 の年式は参考値にとどめること。",
    )

    clean = _clean_listings(listings, model_year_from)
    if not clean:
        return

    years = sorted({r["model_year"] for r in clean})
    cell_prices: dict[tuple[str, int], list[float]] = {}
    for row in clean:
        cell_prices.setdefault((row["vehicle_name"], row["model_year"]), []).append(
            row["price"] / MANYEN
        )

    names = sorted(
        {r["vehicle_name"] for r in clean},
        key=lambda n: -sum(len(v) for (name, _), v in cell_prices.items() if name == n),
    )

    def matrix(start: int, label: str, value_fn, fmt: str, color: bool) -> int:
        ws.cell(row=start - 1, column=1, value=label).font = Font(bold=True, size=11)
        _header(ws, start, ["車種"] + [f"{y}年" for y in years])
        for offset, name in enumerate(names, start=1):
            cell = ws.cell(row=start + offset, column=1, value=name)
            cell.border = BORDER
            for col_offset, year in enumerate(years, start=2):
                value = value_fn(cell_prices.get((name, year), []))
                target = ws.cell(row=start + offset, column=col_offset, value=value)
                target.number_format = fmt
                target.border = BORDER
            if color:
                span = (
                    f"{get_column_letter(2)}{start + offset}:"
                    f"{get_column_letter(1 + len(years))}{start + offset}"
                )
                ws.conditional_formatting.add(span, ROW_SCALE)
        return start + len(names)

    price_last = matrix(
        5,
        "■ 落札中央値（万円）",
        lambda vs: round(_median(vs), 1) if vs else None,
        MANYEN_FMT,
        color=True,
    )
    matrix(
        price_last + 3,
        "■ 落札件数（この数が少ないほど上の中央値はブレる）",
        lambda vs: len(vs) or None,
        INT_FMT,
        color=False,
    )

    # 値落ちカーブ。年式が横軸、車種がシリーズ。
    chart = LineChart()
    chart.title = "年式別の落札中央値（値落ちカーブ）"
    chart.y_axis.title = "万円"
    chart.x_axis.title = "年式"
    chart.height = 13
    chart.width = 30
    chart.add_data(
        Reference(ws, min_col=1, max_col=1 + len(years), min_row=6, max_row=price_last),
        titles_from_data=True,
        from_rows=True,
    )
    chart.set_categories(
        Reference(ws, min_col=2, max_col=1 + len(years), min_row=5, max_row=5)
    )
    for series in chart.series:
        series.marker = Marker(symbol="circle", size=6)
        series.smooth = False
    ws.add_chart(chart, f"{get_column_letter(3 + len(years))}5")

    ws.column_dimensions["A"].width = 18
    for offset in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + offset)].width = 10


# --------------------------------------------------------------- 3. 狙い目


def bargain_sheet(
    ws: Worksheet,
    price_rows: list[dict[str, Any]],
    retail_mileage: dict[str, Any],
    top_n: int = 25,
) -> None:
    _title(
        ws,
        "落札価格と店頭価格の差が大きい順（車種 × 年式）",
        f"小売プレミアム = (小売中央値 / 落札中央値 − 1) × 100。落札件数 {MIN_SAMPLES} 件以上の"
        "年式だけを対象にしている。同じ年式どうしの比較なので年式構成のズレはないが、"
        "**落札車の走行距離**の列を必ず併せて見ること。ヤフオクは個人出品が多く、"
        "距離が伸びた車・保証なし・名義変更や陸送が自己負担のものが混ざる。"
        "差額がそのまま儲けになるわけではない。",
    )

    rows = [
        r
        for r in price_rows
        if (r.get("auction_n") or 0) >= MIN_SAMPLES and r.get("retail_premium_pct") is not None
    ]
    rows.sort(key=lambda r: -r["retail_premium_pct"])
    rows = rows[:top_n]

    start = 4
    _header(
        ws,
        start,
        ["車種・年式", "小売プレミアム\n(%)", "落札中央値\n(万円)", "小売中央値\n(万円)",
         "差額\n(万円)", "落札件数", "掲載台数", "落札車の\n中央走行距離(km)",
         "掲載車の\n中央走行距離(km)\n※車種全体"],
    )

    for offset, row in enumerate(rows, start=1):
        values = [
            f"{row['vehicle_name']} {row['model_year']}年"
            + (f"（{row['generation']}）" if row.get("generation") else ""),
            row["retail_premium_pct"],
            row["auction_median_manyen"],
            row["retail_median_manyen"],
            row["retail_minus_auction_manyen"],
            row["auction_n"],
            row["retail_n"],
            row.get("auction_median_mileage_km"),
            retail_mileage.get(row["vehicle_name"]),
        ]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=start + offset, column=col_index, value=value)
            cell.border = BORDER
            if col_index == 2:
                cell.number_format = PCT_FMT
            elif col_index in (3, 4, 5):
                cell.number_format = MANYEN_FMT
            elif col_index in (6, 7, 8, 9):
                cell.number_format = INT_FMT

    if not rows:
        ws.cell(row=start + 1, column=1, value=f"落札件数 {MIN_SAMPLES} 件以上の年式がまだありません")
        return

    last = start + len(rows)
    chart = BarChart()
    chart.type = "bar"  # 横棒。ラベルが長いので縦棒より読みやすい
    chart.title = "店頭価格 ÷ 落札価格 − 1（%）"
    chart.x_axis.title = "%"
    chart.height = max(10, 0.55 * len(rows))
    chart.width = 26
    chart.add_data(
        Reference(ws, min_col=2, max_col=2, min_row=start, max_row=last),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "K4")

    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[start].height = 32


# ----------------------------------------------------------------- entry


def build(
    wb,
    price_rows: list[dict[str, Any]],
    listings: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
    model_year_from: int,
) -> None:
    retail_mileage = {
        s["vehicle_name"]: s.get("carsensor_retail_median_mileage_km")
        for s in summaries
    }
    bargain_sheet(wb.create_sheet("グラフ_価格差"), price_rows, retail_mileage)
    vehicle_sheet(
        wb.create_sheet("グラフ_車種別"), price_rows, listings, retail_mileage, model_year_from
    )
    year_sheet(wb.create_sheet("グラフ_年式別"), listings, model_year_from)
    log.info("  グラフシート 3 枚を追加")
