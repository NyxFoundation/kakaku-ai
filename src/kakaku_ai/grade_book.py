"""**1 車種の 1 グレードを狙って探す**ための xlsx。

`books.catalog_book()` が「どの車種にするか」を決める本なのに対して、こちらは
車種もグレードも決まっている人が **いま出ている個体を 1 台ずつ見る**ための本。
「2018〜2021年式の A8 60 TFSI を探している」のような状況で使う。

カーセンサーの在庫一覧はグレードを絞る検索軸を持っていないので、車種と年式で
引いてから**出品タイトルでグレードを判定する**。タイトルは
「A8 60 TFSI クワトロ 4WD MMIナビゲーション …」のように先頭にグレードが来る。

同じ車種の他グレードも別シートに残す。狙いのグレードが 8 台しか無いとき、
「55 TFSI なら 15 台あって 100 万安い」が見えるかどうかで判断が変わるため。
"""

from __future__ import annotations

import logging
import re
import statistics as st
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from . import classics, running_cost, store
from .charts import BORDER, INT_FMT, MANYEN_FMT, _header, _title
from .excel import TITLE_FONT, _write_table
from .http import Fetcher
from .sources import carsensor_listings as cl
from .vehicles import DATA_DIR
from .wide import load_catalog

log = logging.getLogger(__name__)

OUTPUT_DIR = DATA_DIR / "xlsx"


def collect_stock(
    fetcher: Fetcher,
    carsensor_code: str,
    model_name: str,
    year_from: int,
    year_to: int,
    *,
    max_pages: int = 10,
) -> list[dict[str, Any]]:
    """その車種の在庫を年式レンジで拾う。グレードの判定はあとで。"""
    vehicle = classics._Vehicle(carsensor_code, model_name)
    rows = cl.fetch_range(fetcher, vehicle, carsensor_code,
                          year_from, year_to, max_pages=max_pages)
    log.info("  %s %s〜%s: %s台", model_name, year_from, year_to, len(rows))
    return rows


def matches_grade(title: str, pattern: str) -> bool:
    """出品タイトルの**先頭のほう**でグレードを判定する。

    タイトルは「A8 60 TFSI クワトロ 4WD Bang&Olufsen/…」のように
    車種名 → グレード → 装備 と並ぶ。装備の羅列まで見に行くと
    「60」が別の意味（60周年記念、160馬力…）で当たるので、
    先頭 24 文字だけを見る。
    """
    return bool(re.search(pattern, (title or "")[:24]))


def build(
    *,
    output: Path,
    title: str,
    carsensor_code: str,
    model_name: str,
    grade_label: str,
    grade_pattern: str,
    year_from: int,
    year_to: int,
    stock: list[dict[str, Any]],
    displacement_l: float,
    assumptions: running_cost.Assumptions | None = None,
) -> Path:
    assumptions = assumptions or running_cost.Assumptions()
    catalog = load_catalog().get(carsensor_code) or {}
    today_year = datetime.now().year

    target = [r for r in stock if matches_grade(r.get("title") or "", grade_pattern)]
    others = [r for r in stock if r not in target]
    for row in stock:
        row["maker"] = catalog.get("maker")
        row["model_name"] = model_name
    classics.rescore(stock)

    # 関連データ。どれも無くても落ちないようにする
    auctions, _ = store.read_latest("yahoo_used_cars")
    auctions = [r for r in auctions if (r.get("model_name") or "") == model_name]
    by_year, _ = store.read_latest("wide_by_year")
    by_year = [r for r in by_year if r.get("carsensor_code") == carsensor_code]
    defects, _ = store.read_latest("catalog_defect_summary")
    defects = [r for r in defects if r.get("carsensor_code") == carsensor_code]
    reviews, _ = store.read_latest("catalog_review_summary")
    reviews = [r for r in reviews if r.get("carsensor_code") == carsensor_code]

    cost_rows = _cost_rows(target, defects, displacement_l, assumptions, today_year)

    wb = Workbook()
    _readme(wb.active, title, grade_label, year_from, year_to,
            target, others, auctions, defects, displacement_l, assumptions)
    wb.active.title = "README"

    if target:
        _price_chart(wb.create_sheet("グラフ_価格と走行"), target, others, grade_label)

    _write_table(wb.create_sheet(f"候補_{grade_label}"), STOCK_COLUMNS,
                 sorted((_flatten(r) for r in target), key=lambda r: -(r["score"] or 0)),
                 number_formats=STOCK_FORMATS, wrap_columns={"why", "title"})
    if others:
        _write_table(wb.create_sheet("他グレード"), STOCK_COLUMNS,
                     sorted((_flatten(r) for r in others),
                            key=lambda r: (r.get("grade_head") or "", -(r["score"] or 0))),
                     number_formats=STOCK_FORMATS, wrap_columns={"why", "title"})
    if cost_rows:
        _write_table(wb.create_sheet("年間維持費"), COST_COLUMNS, cost_rows,
                     number_formats=COST_FORMATS, wrap_columns={"repair_devices"})
    if auctions:
        _write_table(wb.create_sheet("ヤフオク落札"), AUCTION_COLUMNS,
                     sorted((_flatten_auction(r) for r in auctions),
                            key=lambda r: r["end_date"], reverse=True),
                     number_formats=AUCTION_FORMATS, wrap_columns={"title"})
    if by_year:
        _write_table(wb.create_sheet("年式別相場"), YEAR_COLUMNS,
                     sorted(by_year, key=lambda r: r["model_year"]),
                     number_formats=YEAR_FORMATS)
    if defects:
        _write_table(wb.create_sheet("壊れやすい点"), DEFECT_COLUMNS,
                     sorted(defects, key=lambda r: -(r.get("report_count") or 0)),
                     number_formats=DEFECT_FORMATS, wrap_columns={"examples"})
    if reviews:
        _write_table(wb.create_sheet("口コミ"), REVIEW_COLUMNS, reviews,
                     number_formats={"review_count": INT_FMT},
                     wrap_columns={"good_points", "bad_points"})

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    log.info("%s（%s %s台 / 他グレード %s台）",
             output.name, grade_label, len(target), len(others))
    return output


# ------------------------------------------------------------------ 列


STOCK_COLUMNS: list[tuple[str, str]] = [
    ("model_year", "年式"),
    ("grade_head", "グレード"),
    ("score", "状態スコア"),
    ("total_price_manyen", "支払総額\n(万円)"),
    ("base_price_manyen", "車両本体\n(万円)"),
    ("mileage_mankm", "走行距離\n(万km)"),
    ("repair_history", "修復歴"),
    ("inspection", "車検"),
    ("warranty", "保証"),
    ("why", "評価の理由"),
    ("title", "装備（出品タイトル）"),
    ("url", "掲載URL"),
]
STOCK_FORMATS = {
    "model_year": "0", "score": "+0.0;-0.0;0.0",
    "total_price_manyen": MANYEN_FMT, "base_price_manyen": MANYEN_FMT,
    "mileage_mankm": "0.0",
}

COST_COLUMNS: list[tuple[str, str]] = [
    ("model_year", "年式"), ("age", "車齢"),
    ("vehicle_tax_yen", "自動車税"), ("weight_tax_yen", "重量税(年割)"),
    ("compulsory_insurance_yen", "自賠責(年割)"), ("voluntary_insurance_yen", "任意保険"),
    ("inspection_yen", "車検(年割)"), ("maintenance_yen", "整備・消耗品"),
    ("annual_km", "年間走行(km)"), ("fuel_economy_kml", "実燃費(km/L)"),
    ("fuel_yen", "燃料"), ("repair_devices", "保有中に来る故障"),
    ("repair_reserve_yen", "整備予備費/年"),
    ("cash_out_yen", "現金支出/年"), ("monthly_yen", "月あたり"),
]
COST_FORMATS = {
    "model_year": "0", "age": "0", "fuel_economy_kml": "0.0",
    **{k: INT_FMT for k in ("vehicle_tax_yen", "weight_tax_yen",
                            "compulsory_insurance_yen", "voluntary_insurance_yen",
                            "inspection_yen", "maintenance_yen", "annual_km",
                            "fuel_yen", "repair_reserve_yen", "cash_out_yen",
                            "monthly_yen")},
}

AUCTION_COLUMNS: list[tuple[str, str]] = [
    ("end_date", "終了日"), ("model_year", "年式"),
    ("price_manyen", "落札額(万円)"), ("mileage_mankm", "走行(万km)"),
    ("repair_label", "修復歴"), ("bid_count", "入札数"),
    ("seller_label", "出品者"), ("title", "タイトル"), ("url", "URL"),
]
AUCTION_FORMATS = {"model_year": "0", "price_manyen": MANYEN_FMT,
                   "mileage_mankm": "0.0", "bid_count": INT_FMT}

YEAR_COLUMNS: list[tuple[str, str]] = [
    ("model_year", "年式"), ("is_open_bucket", "以前/以降まとめ"),
    ("listing_count", "掲載台数"), ("retail_p25_manyen", "25%(万円)"),
    ("retail_median_manyen", "中央値(万円)"), ("retail_p75_manyen", "75%(万円)"),
    ("url", "URL"),
]
YEAR_FORMATS = {"model_year": "0", "listing_count": INT_FMT,
                **{k: MANYEN_FMT for k in ("retail_p25_manyen",
                                           "retail_median_manyen",
                                           "retail_p75_manyen")}}

DEFECT_COLUMNS: list[tuple[str, str]] = [
    ("defective_device", "不具合装置"), ("report_count", "通報件数"),
    ("share_pct", "構成比(%)"), ("median_mileage_km", "発生時の走行距離\n中央値(km)"),
    ("model_year_min", "対象最古年式"), ("model_year_max", "対象最新年式"),
    ("examples", "実際の症状"),
]
DEFECT_FORMATS = {"report_count": INT_FMT, "share_pct": "0.0",
                  "median_mileage_km": INT_FMT,
                  "model_year_min": "0", "model_year_max": "0"}

REVIEW_COLUMNS: list[tuple[str, str]] = [
    ("review_count", "口コミ件数"), ("score_overall", "総合"),
    ("score_design", "デザイン"), ("score_driving", "走行性能"),
    ("score_ride", "乗り心地"), ("score_price", "価格"),
    ("score_fuel_economy", "燃費"),
    ("good_points", "満足している点"), ("bad_points", "不満な点"),
]


# ------------------------------------------------------------------ 中身


def _grade_head(title: str) -> str:
    """タイトルの先頭から車種名を落として、グレードらしき部分だけ返す。"""
    parts = (title or "").split()
    return " ".join(parts[1:4]) if len(parts) > 1 else ""


def _flatten(row: dict[str, Any]) -> dict[str, Any]:
    out = dict(row)
    out["mileage_mankm"] = round(row["mileage_km"] / 10000, 1) if row.get("mileage_km") else None
    out["why"] = " / ".join(row.get("why") or [])
    out["score"] = round(row.get("score") or 0, 1)
    out["grade_head"] = _grade_head(row.get("title") or "")
    return out


def _flatten_auction(row: dict[str, Any]) -> dict[str, Any]:
    is_store = row.get("seller_is_store")
    return {
        **row,
        "end_date": (row.get("end_time") or "")[:10],
        "price_manyen": round(row["price"] / 10000, 1) if row.get("price") else None,
        "mileage_mankm": (round(row["mileage_km"] / 10000, 1)
                          if row.get("mileage_km") else None),
        "repair_label": {"NONE": "なし", "REPAIRED": "あり",
                         "EXISTS": "あり"}.get(row.get("repair_type") or "", "不明"),
        "seller_label": "不明" if is_store is None else ("業者" if is_store else "個人"),
    }


def _cost_rows(target: list[dict[str, Any]], defects: list[dict[str, Any]],
               displacement_l: float, assumptions: running_cost.Assumptions,
               today_year: int) -> list[dict[str, Any]]:
    """狙いのグレードの年式ごとに維持費を出す。

    走行距離は在庫の実測から。落札が無い（＝この価格帯は個人売買に出てこない）
    車なので、落札側からは取れない。

    **年間走行は年式ごとではなく、対象グレード全体の中央値を使う。** 年式ごとに
    1〜2台しかないと、たまたま車庫保管の個体に当たったときに「年750km」のような
    数字になって燃料費が桁で狂う（実際 A8 2018年式が 6,000km/8年でそうなった）。
    一方、走行距離そのもの（＝いま何km）は年式ごとの中央値でよい。修理の
    見積もりはその個体が何kmかで決まるため。
    """
    by_year: dict[int, list[dict[str, Any]]] = {}
    for row in target:
        if row.get("model_year"):
            by_year.setdefault(row["model_year"], []).append(row)

    per_car = [r["mileage_km"] / max(today_year - r["model_year"], 1)
               for r in target if r.get("mileage_km") and r.get("model_year")]
    typical_km = st.median(per_car) if per_car else assumptions.annual_km

    out = []
    for year, rows in sorted(by_year.items()):
        mileages = [r["mileage_km"] for r in rows if r.get("mileage_km")]
        prices = [r["total_price_manyen"] for r in rows if r.get("total_price_manyen")]
        odometer = st.median(mileages) if mileages else 0
        annual_km = typical_km
        repair = running_cost.repair_outlook(
            defects, odometer_km=odometer, annual_km=annual_km, assumptions=assumptions
        ) if defects else None
        out.append(running_cost.estimate(
            vehicle_name="", model_year=year,
            price_manyen=st.median(prices) if prices else 0,
            displacement_l=displacement_l, depreciation_yen=None,
            depreciation_basis="", assumptions=assumptions,
            annual_km=annual_km, repair=repair, today_year=today_year,
        ))
    return out


def _price_chart(ws, target: list[dict[str, Any]], others: list[dict[str, Any]],
                 grade_label: str) -> None:
    _title(ws, f"{grade_label} の在庫（年式別）と他グレードとの差",
           "狙いのグレードが何台あって、いくらで、どれだけ走っているか。"
           "他グレードとの価格差が大きければ、そちらも選択肢に入る。")

    rows = []
    for label, group in (("狙い: " + grade_label, target), ("他グレード", others)):
        by_year: dict[int, list[dict[str, Any]]] = {}
        for r in group:
            if r.get("model_year"):
                by_year.setdefault(r["model_year"], []).append(r)
        for year, g in sorted(by_year.items()):
            prices = [r["total_price_manyen"] for r in g if r.get("total_price_manyen")]
            km = [r["mileage_km"] for r in g if r.get("mileage_km")]
            rows.append((label, year, len(g),
                         round(min(prices), 1) if prices else None,
                         round(st.median(prices), 1) if prices else None,
                         round(max(prices), 1) if prices else None,
                         round(st.median(km) / 10000, 1) if km else None))

    start = 4
    _header(ws, start, ["区分", "年式", "台数", "最安(万円)", "中央値(万円)",
                        "最高(万円)", "走行中央値(万km)"])
    for offset, record in enumerate(rows):
        for col, value in enumerate(record, start=1):
            cell = ws.cell(row=start + 1 + offset, column=col, value=value)
            cell.border = BORDER
            if col in (4, 5, 6):
                cell.number_format = MANYEN_FMT
            elif col in (3,):
                cell.number_format = INT_FMT
            elif col == 7:
                cell.number_format = "0.0"
    last = start + len(rows)
    if not rows:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"
    chart.title = "年式別の価格帯（万円）"
    chart.x_axis.title = "万円"
    chart.height, chart.width = 12, 24
    chart.add_data(Reference(ws, min_col=4, max_col=6, min_row=start, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=2, min_row=start + 1, max_row=last))
    ws.add_chart(chart, "I4")

    ws.conditional_formatting.add(
        f"E{start + 1}:E{last}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B"),
    )
    ws.column_dimensions["A"].width = 20
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 15


def _readme(ws, title: str, grade_label: str, year_from: int, year_to: int,
            target: list[dict[str, Any]], others: list[dict[str, Any]],
            auctions: list[dict[str, Any]], defects: list[dict[str, Any]],
            displacement_l: float, assumptions: running_cost.Assumptions) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 104
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    prices = [r["total_price_manyen"] for r in target if r.get("total_price_manyen")]
    km = [r["mileage_km"] for r in target if r.get("mileage_km")]
    lines = [
        ("生成日時", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("対象", f"{year_from}〜{year_to}年式 / {grade_label} / 全国"),
        ("狙いのグレード",
         f"{len(target)}台"
         + (f"（{min(prices):.0f}〜{max(prices):.0f}万円 / "
            f"走行 {min(km) / 10000:.1f}〜{max(km) / 10000:.1f}万km）"
            if prices and km else "")),
        ("同年式の他グレード", f"{len(others)}台"),
        ("排気量", f"{displacement_l}L（自動車税・重量税・燃費はこれで決まる）"),
        ("", ""),
        ("■ シート", ""),
        (f"候補_{grade_label}", "狙いのグレードの在庫を1台ずつ。状態スコア順。"
                            "走行距離・修復歴・車検残・保証で採点している。"),
        ("他グレード", "同じ車種・同じ年式の他グレード。"
                  "狙いのグレードが少ないときの比較用。"),
        ("グラフ_価格と走行", "年式ごとの価格帯と走行距離。他グレードとの差も並べてある。"),
        ("年間維持費", "税・保険・車検・燃料・整備予備費の年額。"),
        ("ヤフオク落札", "過去180日の落札実績。**この価格帯の車はほとんど出てこない**"
                    "ので件数は期待できない。"),
        ("年式別相場", "カーセンサー掲載の年式別価格分布（全グレード込み）。"),
        ("壊れやすい点", "国交省の不具合通報を装置別に。発生時の走行距離つき。"),
        ("", ""),
        ("■ グレードの判定について", ""),
        ("やり方",
         "カーセンサーの在庫検索にグレードで絞る軸が無いので、車種と年式で引いてから"
         "**出品タイトルの先頭でグレードを判定**している。タイトルは"
         "「A8 60 TFSI クワトロ 4WD …」のように 車種名→グレード→装備 と並ぶ。"
         "装備の羅列まで見ると「60」が別の意味で当たるので、先頭24文字だけを見ている。"),
        ("取りこぼし",
         "販売店がグレードを書いていない出品は拾えない。「他グレード」シートに"
         "落ちているので、そちらも一度は目で見ること。"),
        ("", ""),
        ("■ 数字の注意", ""),
        ("維持費",
         f"税・自賠責は法定。任意保険{assumptions.voluntary_insurance:,}円・"
         f"車検基本料{assumptions.inspection_fee_2y:,}円/2年・"
         f"整備{assumptions.maintenance:,}円・{assumptions.fuel_yen_per_litre}円/L・"
         "駐車場0円を仮定。輸入車は部品代と工賃が国産より高いので、"
         "**整備費は多めに見ておくこと**。"),
        ("走行距離",
         "在庫の走行距離中央値を車齢で割って年間走行としている。"
         "この価格帯は落札実績がほとんど無いので、落札側からは取れない。"),
        ("整備予備費",
         "国交省の不具合情報にある「装置ごとの発生時走行距離」を保有5年で通過するかで"
         "見積もったもの。**通報件数から発生確率は出せない**ので、通過する装置の"
         "修理費を全部足して 0.5 を掛けている。この 0.5 は素の仮定。"
         f"この車種の通報は全部で{sum(d.get('report_count') or 0 for d in defects)}件で、"
         "母数が小さいぶん装置の並びもぶれる。"),
        ("値落ち",
         "この本には入れていない。年式ごとの落札実績が無く、"
         "掲載価格だけでは1年でいくら落ちるかを出せないため。"),
    ]
    for i, (key, value) in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=key)
        if key.startswith("■"):
            cell.font = TITLE_FONT
        body = ws.cell(row=i, column=2, value=value)
        body.alignment = body.alignment.copy(wrap_text=True, vertical="top")
